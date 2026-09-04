"""Unit tests for the IB-style saved-report Excel model."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import excel_model as em


SAMPLE_MD = """
# SECTION 1 — Executive Summary
Overall rating: **Buy**
12-month price target: **$48.00**
Current Price: $40.00
Implied upside: 20.0%

Acme is the low-cost producer in widgets with a widening moat.

# SECTION 3 — Financial Deep Dive

| Metric | TTM (as of 2025-12-31) | FY2025 (ended 2025-12-31) | FY2024 (ended 2024-12-31) |
|---|---|---|---|
| Revenue ($M) | 1100.0 | 1000.0 | 900.0 |
| Operating Income ($M) | 220.0 | 200.0 | 160.0 |
| Net Income ($M) | 165.0 | 150.0 | 120.0 |
| Diluted EPS | 1.65 | 1.50 | 1.20 |

# SECTION 4 — Growth Analysis

| Metric | FY2026E | FY2027E | FY2028E |
|---|---|---|---|
| Revenue ($M) | 1200.0 | 1320.0 | 1450.0 |
| Operating Income ($M) | 250.0 | 285.0 | 320.0 |
| Net Income ($M) | 190.0 | 215.0 | 245.0 |
| Diluted EPS | 1.90 | 2.15 | 2.45 |
| Free Cash Flow ($M) | 180.0 | 210.0 | 240.0 |

# SECTION 5 — Valuation

WACC 8.5%. Terminal growth 2.5%. Implied share price $47.00.
Enterprise value $9,400m.

| Component | Value | Notes |
|-----------|-------|-------|
| Risk-free rate | 4.0% | 10y UST |
| Equity risk premium | 5.0% | Damodaran |
| Beta (levered) | 1.10 | |
| Cost of equity | 9.5% | rf + β×ERP |
| Pre-tax cost of debt | 5.0% | |
| Tax rate | 21.0% | |
| After-tax cost of debt | 3.95% | |
| E / (D+E) | 90.0% | |
| D / (D+E) | 10.0% | |
| **WACC (base)** | **8.5%** | |

### DCF Projection Ladder (Base Case)

| Year | Fiscal | Revenue ($M) | Rev Growth % | FCF ($M) | FCF Growth % | Discount Factor | PV of FCF ($M) |
|------|--------|--------------|--------------|----------|--------------|-----------------|----------------|
| 0 (A) | FY2025 | 1000.0 | — | 140.0 | — | 1.00 | — |
| 1 (E) | FY2026 | 1200.0 | 20.0% | 180.0 | 28.6% | 0.922 | 166.0 |
| 2 (E) | FY2027 | 1320.0 | 10.0% | 210.0 | 16.7% | 0.849 | 178.3 |
| 3 (E) | FY2028 | 1450.0 | 9.8% | 240.0 | 14.3% | 0.783 | 187.9 |
| 4 (E) | FY2029 | 1595.0 | 10.0% | 260.0 | 8.3% | 0.722 | 187.7 |
| 5 (E) | FY2030 | 1754.5 | 10.0% | 280.0 | 7.7% | 0.665 | 186.2 |
| **Sum PV explicit FCF** | — | — | — | — | — | — | **906.1** |

| Step | Amount ($M) | Formula / note |
|------|-------------|----------------|
| Year-5 FCF | 280.0 | from ladder |
| Terminal growth (g) | 2.5% | GDP |
| Terminal value (exit) | 4783.3 | FCF5×(1+g)/(WACC−g) |
| PV of terminal value | 3181.0 | TV / (1+WACC)^5 |
| Enterprise value | 4087.1 | Σ PV FCF + PV(TV) |
| (−) Net debt | 120.0 | |
| Equity value | 3967.1 | |
| Diluted shares (M) | 100.0 | |
| **DCF value / share** | **$39.67** | |

| DCF Sensitivity | TGR: 1.0% | TGR: 2.0% | TGR: 2.5% | TGR: 3.0% | TGR: 3.5% |
|-----------------|-----------|-----------|-----------|-----------|-----------|
| WACC: 7.0% | $48.00 | $52.00 | $55.00 | $58.00 | $62.00 |
| WACC: 8.0% | $42.00 | $45.00 | $47.00 | $50.00 | $53.00 |
| WACC: 8.5% | $39.00 | $42.00 | **$44.00** | $47.00 | $50.00 |
| WACC: 9.0% | $36.00 | $39.00 | $41.00 | $44.00 | $47.00 |
| WACC: 10.0% | $31.00 | $33.00 | $35.00 | $37.00 | $40.00 |

| Method | Implied Value | Weight | Weighted Value |
|---|---|---|---|
| DCF (base) | $47.00 | 60% | $28.20 |
| DCF (7%/3.5% bull duration) | $80.00 | 0% | $0.00 |
| EV/Revenue | $55.00 | 15% | $8.25 |
| P/E | $52.00 | 15% | $7.80 |
| Street-anchored | $50.00 | 10% | $5.00 |
| **12M Price Target** | — | — | **$48.00** |

| Ticker | P/E | EV/EBITDA | FCF yield |
|---|---|---|---|
| ACME | 24.0x | 14.0x | 4.5% |
| PEER | 22.0x | 13.0x | 5.0% |

# SECTION 7.5 — Institutional Analyst Consensus

| Firm | Rating | 12M Price Target | Upside vs Current | Date | Action |
|---|---|---|---|---|---|
| Goldman Sachs | Buy | $52.00 | 30.0% | 2026-01-15 | Maintain |
| Morgan Stanley | Overweight | $49.00 | 22.5% | 2026-01-10 | Raise |

# SECTION 8 — The Verdict
Bull case: Price target $62 (30% probability)
Base case: Price target $48 (50% probability)
Bear case: Price target $28 (20% probability)
"""


def _financials():
    return {
        "ticker": "ACME",
        "entity_name": "Acme Widget Corp",
        "source": "company_financials_db",
        "annuals": [
            {
                "fy": 2025, "end": "2025-12-31",
                "Revenue": 1_000_000_000, "OperatingIncome": 200_000_000,
                "NetIncome": 150_000_000, "DilutedEPS": 1.50,
                "DilutedShares": 100_000_000, "FreeCashFlow": 140_000_000,
                "Cash": 80_000_000, "TotalDebt": 200_000_000,
                "StockholdersEquity": 500_000_000, "TotalAssets": 900_000_000,
                "EBITDA": 250_000_000, "OperatingCashFlow": 180_000_000,
                "CapEx": -40_000_000, "GrossProfit": 400_000_000,
                "GrossMargin": 0.40, "OperatingMargin": 0.20, "NetMargin": 0.15,
            },
            {
                "fy": 2024, "end": "2024-12-31",
                "Revenue": 900_000_000, "OperatingIncome": 160_000_000,
                "NetIncome": 120_000_000, "DilutedEPS": 1.20,
                "DilutedShares": 100_000_000, "FreeCashFlow": 110_000_000,
                "Cash": 60_000_000, "TotalDebt": 210_000_000,
                "StockholdersEquity": 420_000_000, "TotalAssets": 820_000_000,
                "EBITDA": 210_000_000, "OperatingCashFlow": 150_000_000,
                "CapEx": -40_000_000, "GrossProfit": 350_000_000,
            },
        ],
        "ttm": {
            "end": "2025-12-31", "method": "bridge",
            "Revenue": 1_100_000_000, "NetIncome": 165_000_000,
            "OperatingIncome": 220_000_000, "FreeCashFlow": 155_000_000,
            "DilutedEPS": 1.65, "Cash": 80_000_000, "TotalDebt": 200_000_000,
        },
        "quarterly": {
            "current": {
                "fy": 2025, "fp": "Q4", "end": "2025-12-31",
                "Revenue": 280_000_000, "NetIncome": 42_000_000, "DilutedEPS": 0.42,
            },
            "prior_year_same_q": {
                "fy": 2024, "fp": "Q4", "end": "2024-12-31",
                "Revenue": 250_000_000, "NetIncome": 35_000_000, "DilutedEPS": 0.35,
            },
        },
    }


def test_parse_cell_number():
    assert em.parse_cell_number("$1,234.5") == 1234.5
    assert em.parse_cell_number("(123.4)") == -123.4
    assert em.parse_cell_number("12.5%") == 0.125
    assert em.parse_cell_number("N/A") is None
    assert em.parse_cell_number("—") is None
    assert em.parse_cell_number("$9.4bn") == 9400.0
    assert em.parse_cell_number("$47.5 billion") == 47500.0
    assert em.parse_cell_number("13.3%") == 0.133


def test_parse_tables_and_forecasts():
    tables = em.parse_md_tables(SAMPLE_MD)
    assert len(tables) >= 4
    fc = em.extract_forecasts(tables)
    assert 2026 in fc and 2027 in fc
    assert fc[2026]["Revenue"] == 1200.0
    assert fc[2026]["DilutedEPS"] == 1.90
    street = em.extract_street_table(tables)
    assert street and street["rows"][0][0].startswith("Goldman")


def test_dcf_and_scenarios():
    dcf = em.extract_dcf(SAMPLE_MD)
    assert abs(dcf["wacc"] - 0.085) < 1e-9
    assert abs(dcf["terminal_growth"] - 0.025) < 1e-9
    assert dcf["implied_price"] == 47.0
    sc = em.extract_scenarios(SAMPLE_MD)
    cases = {r["case"]: r for r in sc}
    assert cases["Bull"]["price_target"] == 62.0
    assert cases["Base"]["price_target"] == 48.0
    assert abs(cases["Bull"]["probability"] - 0.30) < 1e-9


def test_wacc_ladder_sensitivity_parsers():
    tables = em.parse_md_tables(SAMPLE_MD)
    wacc = em.extract_wacc_build(tables)
    assert abs(wacc["rf"] - 0.04) < 1e-9
    assert abs(wacc["erp"] - 0.05) < 1e-9
    assert abs(wacc["beta"] - 1.10) < 1e-9
    assert abs(wacc["wacc"] - 0.085) < 1e-9
    assert abs(wacc["we"] - 0.90) < 1e-9
    ladder = em.extract_dcf_ladder(tables)
    by_t = {r["t"]: r for r in ladder}
    assert by_t[1]["fcf"] == 180.0
    assert by_t[5]["fcf"] == 280.0
    assert by_t[1]["fy"] == 2026
    sens = em.extract_sensitivity(tables)
    assert sens is not None
    assert 0.025 in [round(g, 4) for g in sens["tgrs"]]
    assert round(sens["cells"][(0.085, 0.025)], 2) == 44.0


def test_workbook_sheets_and_pro_forma(tmp_path):
    import openpyxl
    from io import BytesIO

    raw = em.build_ib_model_bytes(
        "ACME",
        financials=_financials(),
        report_md=SAMPLE_MD,
        summary={
            "rating": "Buy",
            "price_target": 48.0,
            "current_price": 40.0,
            "upside_pct": 20.0,
            "thesis": "Low-cost producer with a widening moat.",
            "sector": "Industrials",
        },
        quote={"price": 40.0},
        engine="Rock",
        entity_name="Acme Widget Corp",
        source="unit test",
        dropbox_note="Dropbox: /Reports/ACME_DGA_Model.xlsx",
        generated_at="2026-09-02",
    )
    assert raw[:2] == b"PK"
    wb = openpyxl.load_workbook(BytesIO(raw), data_only=False)
    names = set(wb.sheetnames)
    assert names >= {
        "Cover", "Financial Model", "Valuation", "Scenarios", "Street", "Quarterly",
    }
    model = wb["Financial Model"]
    headers = [model.cell(5, c).value for c in range(1, 12) if model.cell(5, c).value]
    assert "FY2024A" in headers
    assert "FY2025A" in headers
    assert any(str(h).startswith("FY2026E") for h in headers)
    # Revenue actuals are in $ millions
    # Find revenue row
    rev_row = None
    for r in range(1, 40):
        if model.cell(r, 1).value == "Revenue":
            rev_row = r
            break
    assert rev_row
    # FY2024A is first year col (col 2)
    assert model.cell(rev_row, 2).value == 900.0
    # Estimate year present and blue-input
    fy26_col = headers.index("FY2026E") + 1
    assert model.cell(rev_row, fy26_col).value == 1200.0
    cover = wb["Cover"]
    assert "DGA CAPITAL" in str(cover["A1"].value)
    assert cover["A7"].value == "Buy"
    val = wb["Valuation"]
    labels = {val.cell(r, 1).value: r for r in range(1, 80)}
    assert "WACC (base)" in labels
    assert val.cell(labels["WACC (base)"], 2).value == "=B9*B13+B12*B14"
    assert "Risk-free rate" in labels
    assert abs(val.cell(labels["Risk-free rate"], 2).value - 0.04) < 1e-9
    assert "Terminal growth (g)" in labels
    assert "DCF value / share" in labels
    assert val.cell(3, 1).value == "vs last →"
    assert "DCF" in str(val.cell(3, 2).value or "")
    v3 = str(val.cell(3, 3).value or "")
    assert "UNDERVALUED" in v3 and "OVERVALUED" in v3 and "FAIR VALUED" in v3
    # Row-3 verdict cells use dark ink on a light fill (never white-on-white).
    c3_font = (val.cell(3, 3).font.color.rgb or "") if val.cell(3, 3).font and val.cell(3, 3).font.color else ""
    c3_fill = (val.cell(3, 3).fill.fgColor.rgb or "") if val.cell(3, 3).fill and val.cell(3, 3).fill.fgColor else ""
    assert "FFFFFF" not in str(c3_font).upper() or "E2E8F0" in str(c3_fill).upper()
    assert "DCF verdict (vs last)" in labels
    assert "Mispricing vs last" in labels
    assert "$ / sh vs last" in labels
    # DCF (base) reverse bridge sits in columns E–G next to the Gordon walk
    base_title_r = None
    for r in range(20, 90):
        if "DCF (BASE) BRIDGE" in str(val.cell(r, 5).value or ""):
            base_title_r = r
            break
    assert base_title_r, "missing DCF (base) reverse bridge"
    assert val.cell(base_title_r + 1, 5).value == "Step"
    assert val.cell(base_title_r + 2, 5).value == "DCF (base) $/share"
    assert abs(float(val.cell(base_title_r + 2, 6).value) - 47.0) < 1e-9
    gap_labs = [str(val.cell(r, 5).value or "") for r in range(base_title_r, 95)]
    assert any("GAP" in x and "DCF (base)" in x for x in gap_labs)
    assert "Model DCF value / share" in gap_labs
    assert "$ gap (model − base)" in gap_labs
    assert "% gap" in gap_labs
    # reverse walk uses the same shares / net debt / ladder as the model
    assert any(
        str(val.cell(r, 6).value or "") == "=$B$20"
        for r in range(base_title_r, base_title_r + 20)
    )
    assert any(
        str(val.cell(r, 6).value or "").startswith("=H")
        for r in range(base_title_r, base_title_r + 20)
    )
    # Pro forma years header includes FY2026E
    pf_headers = [val.cell(5, c).value for c in range(5, 14)]
    assert "FY2026E" in pf_headers
    assert "FY2025A" in pf_headers
    # Ladder FCF for year 1 is linked to the pro forma FCF row
    assert any(
        isinstance(val.cell(r, 1).value, str) and val.cell(r, 1).value.startswith("1 (E)")
        for r in range(20, 45)
    )
    # Live sensitivity grid uses WACC × g formulas
    found_sens = False
    for r in range(20, 50):
        lab = str(val.cell(r, 10).value or "")
        if lab.startswith("WACC") and "g" in lab and "share" not in lab.lower():
            found_sens = True
            # first data row is a WACC rate with a formula in K
            k = val.cell(r + 1, 11).value
            assert isinstance(k, str) and k.startswith("="), k
            break
    assert found_sens
    street = wb["Street"]
    blob = " ".join(
        str(street.cell(r, c).value or "")
        for r in range(1, 20) for c in range(1, 6)
    )
    assert "Goldman" in blob
    sc = wb["Scenarios"]
    cases = [sc.cell(r, 1).value for r in range(5, 10)]
    assert "Bull" in cases and "Base" in cases and "Bear" in cases


def test_classify_stock_style():
    st = em.classify_stock_style(
        SAMPLE_MD, summary={"current_price": 40.0},
    )
    # DCF ~$47 vs $40 = cheap; fwd rev 1000→1200 = +20% → GARP
    assert st["style"] == "garp"
    assert st["label"] == "GARP"
    assert st["dcf_gap"] is not None and st["dcf_gap"] > 0.05
    assert st["fwd_rev_growth"] is not None and st["fwd_rev_growth"] >= 0.15
    cheap = em.style_from_metrics(50.0, 40.0, 0.05, 0.04)
    assert cheap["style"] == "value"
    grow = em.style_from_metrics(40.0, 40.0, 0.25, None)
    assert grow["style"] == "growth"
    rich = em.style_from_metrics(30.0, 40.0, 0.02, None)
    assert rich["style"] == "expensive" and rich["label"] == "RICH"
    core = em.style_from_metrics(40.0, 40.0, 0.04, None)
    assert core["style"] == "core"
    empty = em.style_from_metrics(None, 40.0, None, None)
    assert empty["style"] is None


def test_inject_cover_dcf_target():
    cover = """| DGA CAPITAL RESEARCH | August 28, 2026 |
|-----------------------------------------------|-----------------------------|
| **CIENA CORP** (CIEN: NYSE) | Equity Research |
| **Rating:** BUY | |
| **12-Month Price Target:** $485 | Implied Return: +28.2% |
| **Current Price:** $378.44 | 52-Week Range: $90–$637 |

# SECTION 7
| **DCF value / share** | **$181.35** |
"""
    out = em.inject_cover_dcf_target(cover)
    assert "DCF Target" in out
    assert "$181.35" in out
    # blended PT row must stay
    assert "**12-Month Price Target:** $485" in out
    # idempotent
    again = em.inject_cover_dcf_target(out)
    assert again.count("DCF Target") == 1


def test_inject_cover_dcf_from_derivation_table_not_pt():
    """NFLX-style: no 'DCF value / share' line; DCF lives in the method table."""
    md = """| DGA CAPITAL RESEARCH | 2026-07-17 |
|-----------------------------------------------|-----------------------------|
| **NETFLIX INC** (NFLX: NASDAQ) | Equity Research |
| **Rating:** BUY                               |                             |
| **12-Month Price Target:** $100               | Implied Return: +45.1%      |
| **Current Price:** $68.95 | 52-Week Range: $70–$127 |

| Method         | Implied Value | Weight | Weighted Value |
|----------------|-------------|--------|---------------|
| DCF            | $110        | 50%    | $55           |
| EV/EBITDA Comps | $95–105 | 30% | $30          |
| **12M Price Target** | —    | —      | **$100**      |
"""
    out = em.inject_cover_dcf_target(md)
    rating = next(ln for ln in out.splitlines() if "Rating:" in ln)
    assert "DCF Target" in rating
    assert "$110" in rating
    assert "**12-Month Price Target:** $100" in out
    assert out.count("DCF Target") == 1


def test_rank_dcf_undervalued():
    rows = em.rank_dcf_undervalued([
        {"ticker": "CHEAP", "dcf_value": 50, "price": 40},   # +25%
        {"ticker": "MEH", "dcf_value": 41, "price": 40},     # +2.5%
        {"ticker": "RICH", "dcf_value": 30, "price": 40},    # overvalued
        {"ticker": "FAIR", "dcf_value": 40, "price": 40},    # flat
        {"ticker": "NO", "dcf_value": None, "price": 40},
        {"ticker": "BEST", "dcf_value": 80, "price": 40},    # +100%
    ], limit=10)
    assert [r["ticker"] for r in rows] == ["BEST", "CHEAP", "MEH"]
    assert rows[0]["dcf_gap_pct"] == 100.0
    top2 = em.rank_dcf_undervalued(
        [{"ticker": "A", "dcf_value": 12, "price": 10},
         {"ticker": "B", "dcf_value": 15, "price": 10},
         {"ticker": "C", "dcf_value": 11, "price": 10}],
        limit=2,
    )
    assert [r["ticker"] for r in top2] == ["B", "A"]


def test_extract_valuation_approaches_not_blended():
    apps = em.extract_valuation_approaches(
        SAMPLE_MD, last=40.0, pt=48.0,
    )
    by_id = {a["id"]: a for a in apps}
    names = [a["name"] for a in apps]
    assert "dcf" in by_id
    assert "dcf_base" in by_id
    assert "ev_rev" in by_id
    assert "pe" in by_id
    assert "street_anchored" in by_id
    assert "street" in by_id
    assert "report_pt" in by_id
    # Scenario bull/base/bear are not chips — "Base case" collided with DCF (base)
    assert "base" not in by_id and "bull" not in by_id and "bear" not in by_id
    assert not any("base case" in (a.get("name") or "").lower() for a in apps)
    # Extra DCF duration case is dropped
    assert not any("bull duration" in (n or "").lower() for n in names)
    assert not any("7%" in (n or "") for n in names)
    # Model DCF value/share from the bridge, not the $47 implied-share prose
    assert abs(by_id["dcf"]["value"] - 39.67) < 0.02
    assert "model" in (by_id["dcf"]["note"] or "").lower()
    # DCF (base) uses implied $47, NOT weighted $28.20
    assert abs(by_id["dcf_base"]["value"] - 47.0) < 1e-9
    assert "implied" in (by_id["dcf_base"]["note"] or "").lower()
    # Multiples / street-anchored: implied, not weighted contribution
    assert abs(by_id["ev_rev"]["value"] - 55.0) < 1e-9
    assert abs(by_id["pe"]["value"] - 52.0) < 1e-9
    assert abs(by_id["street_anchored"]["value"] - 50.0) < 1e-9
    # 12m PT is the only weighted blend
    assert abs(by_id["report_pt"]["value"] - 48.0) < 1e-9
    assert "blend" in (by_id["report_pt"]["note"] or "").lower()
    dcf = by_id["dcf"]
    assert dcf["verdict"] == "FAIR VALUED" or dcf["tone"] in ("fair", "under")
    recut = em.recut_approaches_vs_last([by_id["dcf_base"]], 80.0)
    assert recut[0]["verdict"] == "OVERVALUED"
    fair = em.valuation_verdict(41.0, 40.0)
    assert fair["verdict"] == "FAIR VALUED" and fair["tone"] == "fair"
    fill, ink = em.verdict_palette("under", 0.2)
    assert ink.upper() != "FFFFFF"
    fill_x, ink_x = em.verdict_palette("under", 1.0)
    assert fill_x.upper() != "FFFFFF" or ink_x.upper() != "FFFFFF"


def test_workbook_approaches_table():
    import openpyxl
    from io import BytesIO
    raw = em.build_ib_model_bytes(
        "ACME",
        financials=_financials(),
        report_md=SAMPLE_MD,
        summary={"rating": "Buy", "price_target": 48.0, "current_price": 40.0},
        quote={"price": 40.0},
    )
    wb = openpyxl.load_workbook(BytesIO(raw), data_only=False)
    val = wb["Valuation"]
    blob = " ".join(
        str(val.cell(r, c).value or "")
        for r in range(1, 90) for c in range(1, 7)
    )
    assert "APPROACHES vs LAST" in blob
    assert "DCF (base)" in blob
    assert "DCF value / share" in blob
    assert "EV/Revenue" in blob
    assert "Street consensus" in blob
    # Extra DCF duration case must not appear in the Approaches table
    # (it can still sit in the raw derivation dump further down).
    ap_names = []
    in_ap = False
    for r in range(1, 90):
        a1 = str(val.cell(r, 1).value or "")
        if "APPROACHES vs LAST" in a1:
            in_ap = True
            continue
        if in_ap and (
            a1.startswith("TRADING") or "PRICE TARGET DERIVATION" in a1.upper()
            or a1.startswith("COMPARABLE")
        ):
            break
        if in_ap:
            ap_names.append(a1)
    assert not any("bull duration" in n.lower() or "7%" in n for n in ap_names)
    assert "not a weighted blend" in blob.lower() or "not a weighted" in blob.lower()
    # EV/Revenue uses implied $55, not weighted $8.25
    ev_r = None
    for r in range(1, 90):
        if val.cell(r, 1).value == "EV/Revenue":
            ev_r = r
            break
    assert ev_r
    assert abs(float(val.cell(ev_r, 2).value) - 55.0) < 1e-9
    verd = val.cell(ev_r, 5)
    assert verd.value == "UNDERVALUED"
    fill = str(getattr(verd.fill.fgColor, "rgb", "") or "")
    ink = str(getattr(verd.font.color, "rgb", "") or "")
    assert fill
    assert not (ink.endswith("FFFFFF") and fill.endswith("FFFFFF"))
    assert not (ink.endswith("FFFFFF") and fill.endswith("FBF6E8"))


def test_dcf_verdict_formulas():
    f = em._dcf_verdict_formulas("B44", "$B$21")
    assert f["label"].startswith("=")
    assert "UNDERVALUED" in f["label"]
    assert "OVERVALUED" in f["label"]
    assert "FAIR VALUED" in f["label"]
    assert "0.05" in f["label"]
    assert "cheap" in f["byline"] and "rich" in f["byline"]


def test_to_model_units():
    assert em.to_model_units("Revenue", 1_000_000_000) == 1000.0
    assert em.to_model_units("DilutedEPS", 1.5) == 1.5
    assert em.to_model_units("DilutedShares", 100_000_000) == 100.0
    assert em.to_model_units("GrossMargin", 0.4) == 0.4
    assert em.to_model_units("Revenue", 1200.0, already_millions=True) == 1200.0


def test_growth_label_is_not_revenue():
    md = """
| Metric | FY2026E |
|---|---|
| Revenue growth | 13.3 |
| Diluted EPS | 26.00 |
"""
    fc = em.extract_forecasts(em.parse_md_tables(md))
    assert 2026 in fc
    assert "Revenue" not in fc[2026]
    assert abs(fc[2026]["RevenueGrowth"] - 0.133) < 1e-9


def test_revenue_billions_header_scales_to_millions():
    md = """
| Metric | FY2026E |
|---|---|
| Revenue ($B) | 47.5 |
"""
    fc = em.extract_forecasts(em.parse_md_tables(md))
    assert abs(fc[2026]["Revenue"] - 47500.0) < 1e-6


def test_normalize_shares_prefers_sec_over_three_billion():
    assert em.normalize_shares_millions(3.0, 427.0) == 427.0
    assert abs(em.normalize_shares_millions(0.427, 427.0) - 427.0) < 1.0
    assert abs(em.normalize_shares_millions(427.0, 427.0) - 427.0) < 1e-9
    assert abs(em.normalize_shares_millions(427_000_000, 427.0) - 427.0) < 1e-6


def test_nflx_like_units_do_not_break_valuation():
    """13.3 growth vs $45,183m actual; 3.0 'billion' shares vs 427m SEC."""
    import openpyxl
    from io import BytesIO

    md = """
# SECTION 4 — Growth

| Metric | FY2026E |
|---|---|
| Revenue growth | 13.3 |
| Diluted EPS | 26.00 |

# SECTION 5 — Valuation

WACC 8.5%. Terminal growth 2.5%. diluted shares outstanding 3.0 billion.

| Step | Amount ($M) | Formula / note |
|------|-------------|----------------|
| Diluted shares (M) | 3.0 | |
| **DCF value / share** | **$90.00** | |
"""
    financials = {
        "ticker": "NFLX",
        "entity_name": "Netflix Inc",
        "annuals": [{
            "fy": 2025, "end": "2025-12-31",
            "Revenue": 45_183_000_000,
            "OperatingIncome": 10_000_000_000,
            "NetIncome": 8_000_000_000,
            "DilutedEPS": 18.0,
            "DilutedShares": 427_000_000,
            "FreeCashFlow": 7_000_000_000,
            "Cash": 7_000_000_000,
            "TotalDebt": 14_000_000_000,
            "EBITDA": 12_000_000_000,
            "GrossProfit": 20_000_000_000,
        }],
        "ttm": {
            "end": "2025-12-31",
            "Revenue": 45_183_000_000,
            "DilutedShares": 427_000_000,
            "Cash": 7_000_000_000,
            "TotalDebt": 14_000_000_000,
            "FreeCashFlow": 7_000_000_000,
        },
    }
    raw = em.build_ib_model_bytes(
        "NFLX",
        financials=financials,
        report_md=md,
        summary={"rating": "Buy", "price_target": 90.0, "current_price": 80.0},
        quote={"price": 80.0},
    )
    wb = openpyxl.load_workbook(BytesIO(raw), data_only=False)
    model = wb["Financial Model"]
    headers = [model.cell(5, c).value for c in range(1, 12) if model.cell(5, c).value]
    rev_row = next(r for r in range(1, 40) if model.cell(r, 1).value == "Revenue")
    fy25 = headers.index("FY2025A") + 1
    fy26 = headers.index("FY2026E") + 1
    assert abs(model.cell(rev_row, fy25).value - 45183.0) < 1.0
    fy26_rev = model.cell(rev_row, fy26).value
    assert fy26_rev is not None
    assert abs(fy26_rev - 13.3) > 100
    assert 48_000 < fy26_rev < 55_000
    val = wb["Valuation"]
    share_rows = [r for r in range(1, 80) if val.cell(r, 1).value == "Diluted shares (m)"]
    assert share_rows
    shares = val.cell(share_rows[0], 2).value
    assert shares is not None
    assert abs(float(shares) - 427.0) < 2.0
    # Pro forma year-0 revenue is the actual, not 13.3
    pf = [val.cell(5, c).value for c in range(5, 14)]
    a_col = pf.index("FY2025A") + 5
    e_col = pf.index("FY2026E") + 5
    assert abs(val.cell(6, a_col).value - 45183.0) < 1.0
    e_rev = val.cell(6, e_col).value
    if isinstance(e_rev, (int, float)):
        assert abs(e_rev - 13.3) > 100
        assert e_rev > 40_000
    else:
        assert isinstance(e_rev, str) and e_rev.startswith("=")
