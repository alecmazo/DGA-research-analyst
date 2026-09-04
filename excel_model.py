"""
IB-style Excel model for a saved DGA research report.

Produces a Goldman / MS / BofA-style workbook:

  Cover            — rating, 12m PT, capital structure, thesis
  Financial Model  — historicals (A) + TTM + pro forma (E), 3-statement
  Valuation        — WACC build, pro forma years, DCF ladder, TV/equity
                     bridge, live WACC×g sensitivity, comps / PT weights
  Scenarios        — bull / base / bear + expected value
  Street           — sell-side consensus table from the report
  Quarterly        — last eight reported quarters

Historicals come from company_financials (dollars). Forward / pro forma
numbers and valuation assumptions are parsed from the saved report
markdown. Estimate cells are Excel-blue; actuals are black.

Public API
----------
    path = write_ib_model(ticker, output_path, *, financials, report_md, ...)
    raw  = build_ib_model_bytes(...)
"""

from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

# ── Colors (DGA navy + classic IB worksheet) ────────────────────────────────
NAVY = "0A1628"
GOLD = "C9A84C"
WHITE = "FFFFFF"
INK = "0F172A"
MUTED = "64748B"
BLUE_EST = "0070C0"       # forecast / estimate font
YELLOW = "FFF2CC"         # input cells
SECTION = "1E3A5F"
ROW_ALT = "F8FAFC"
LINE = "CBD5E1"
GREEN_POS = "166534"
RED_NEG = "B91C1C"
PALE_GOLD = "FBF6E8"
PALE_NAVY = "E8EEF5"


# ── Markdown / number parsing ───────────────────────────────────────────────
_NUM_RE = re.compile(
    r"^\s*\(?\s*\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?|[0-9]*\.[0-9]+|[0-9]+)"
    r"\s*(%|x|×|bn|billions?|mm|millions?|b|m|k|thousands?)?\s*\)?\s*$",
    re.I,
)
_YEAR_RE = re.compile(r"(20\d{2})")
_FY2_RE = re.compile(r"FY\s*'?(\d{2})\b", re.I)


def _strip_md(cell: str) -> str:
    s = (cell or "").strip()
    s = s.replace("\u00a0", " ").replace("**", "").replace("__", "")
    s = re.sub(r"</?[^>]+>", "", s)
    s = s.strip(" *")
    return s


def parse_cell_number(raw: str) -> Optional[float]:
    """Parse an IB-table cell into a float.

    Percents become fractions (12.5% → 0.125). Trailing x is kept as the
    multiple. $ / commas / accounting (123.4) negatives are handled.
    """
    s = _strip_md(raw)
    if not s or s in ("—", "–", "-", "N/A", "n/a", "NA", "nm", "NM", "."):
        return None
    neg = s.startswith("(") and s.endswith(")")
    if s.startswith("-") or s.startswith("−"):
        neg = True
        s = s.lstrip("-−")
    m = _NUM_RE.match(s.replace("−", "-"))
    if not m:
        return None
    try:
        val = float(m.group(1).replace(",", ""))
    except ValueError:
        return None
    unit = (m.group(2) or "").lower()
    if unit == "%":
        val = val / 100.0
    else:
        val = _apply_money_unit(val, unit)
    if neg:
        val = -abs(val)
    return val


def _apply_money_unit(val: float, unit: str) -> float:
    """Convert a trailing unit into the model’s $ millions (shares also millions)."""
    u = re.sub(r"s$", "", (unit or "").lower().strip())
    if u in ("bn", "b", "billion"):
        return val * 1000.0
    if u in ("k", "thousand"):
        return val / 1000.0
    return val


def _money_scale_from_text(text: str) -> Optional[float]:
    """Header/label unit → multiplier into $ millions. None if unspecified."""
    s = _strip_md(text or "")
    if re.search(r"billion|\(\s*\$?\s*bn\s*\)|\$bn\b|\(\s*\$b\s*\)", s, re.I):
        return 1000.0
    if re.search(r"million|\(\s*\$?\s*mm?\s*\)|\$mm\b|\$m\b|\(\s*\$m\s*\)", s, re.I):
        return 1.0
    if re.search(r"thousand|\(\s*\$k\s*\)|\$k\b", s, re.I):
        return 0.001
    return None


def _cell_has_explicit_unit(raw: str) -> bool:
    s = _strip_md(raw or "")
    return bool(re.search(
        r"(bn|billions?|mm|millions?|(?<![a-z])b(?![a-z])|(?<![a-z])m(?![a-z])|k|thousands?)\s*$",
        s,
        re.I,
    ))


def parse_md_tables(md: str) -> list[dict[str, Any]]:
    """Return GitHub-flavored markdown tables with a nearby heading."""
    if not md:
        return []
    lines = md.replace("\r\n", "\n").split("\n")
    tables: list[dict[str, Any]] = []
    heading = ""
    i = 0
    n = len(lines)
    while i < n:
        line = lines[i]
        hm = re.match(r"^\s{0,3}#{1,4}\s+(.+?)\s*$", line)
        if hm:
            heading = _strip_md(hm.group(1))
            i += 1
            continue
        if "|" not in line:
            i += 1
            continue
        block = [line]
        j = i + 1
        while j < n and "|" in lines[j]:
            block.append(lines[j])
            j += 1
        parsed = _table_from_lines(block)
        if parsed and len(parsed["headers"]) >= 2 and parsed["rows"]:
            parsed["title"] = heading
            tables.append(parsed)
        i = j
    return tables


def _split_row(line: str) -> list[str]:
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    return [_strip_md(c) for c in s.split("|")]


def _is_sep_row(cells: list[str]) -> bool:
    if not cells:
        return False
    return all(re.fullmatch(r":?-{3,}:?", c.replace(" ", "")) or not c for c in cells)


def _table_from_lines(block: list[str]) -> Optional[dict[str, Any]]:
    rows = [_split_row(ln) for ln in block if ln.strip()]
    rows = [r for r in rows if any(c for c in r)]
    if len(rows) < 2:
        return None
    headers = rows[0]
    body = rows[1:]
    if body and _is_sep_row(body[0]):
        body = body[1:]
    width = len(headers)
    body = [(r + [""] * width)[:width] for r in body]
    if not body:
        return None
    return {"headers": headers, "rows": body}


def _col_is_estimate(header: str) -> bool:
    h = (header or "").strip()
    if re.search(r"(20\d{2}|FY\s*'?\d{2,4})\s*E\b", h, re.I):
        return True
    if re.search(r"\b(est\.?|fwd|forward|ntm|n?tm|proj(?:ection|ected)?|pf|pro\s*forma)\b", h, re.I):
        return True
    return False


def _col_is_ttm(header: str) -> bool:
    return bool(re.search(r"\b(ttm|ltm)\b", header or "", re.I))


def _year_from_header(header: str) -> Optional[int]:
    h = header or ""
    m = _YEAR_RE.search(h)
    if m:
        return int(m.group(1))
    m = _FY2_RE.search(h)
    if m:
        return 2000 + int(m.group(1))
    return None


_METRIC_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"revenue.{0,24}(growth|yoy)|^(total\s+)?sales.{0,16}(growth|yoy)", re.I), "RevenueGrowth"),
    (re.compile(r"(?:free cash flow|fcf).{0,16}(growth|yoy)", re.I), "FCFGrowth"),
    (re.compile(r"(?:diluted\s+)?eps.{0,16}(growth|yoy)", re.I), "EPSGrowth"),
    (re.compile(r"^(total\s+)?revenue|^net sales|^sales\b|^revenues", re.I), "Revenue"),
    (re.compile(r"cost of (revenue|goods|sales)|^cogs", re.I), "CostOfRevenue"),
    (re.compile(r"^gross profit", re.I), "GrossProfit"),
    (re.compile(r"^gross margin", re.I), "GrossMargin"),
    (re.compile(r"operating income|^ebit\b|^op(?:erating)?\.?\s*inc", re.I), "OperatingIncome"),
    (re.compile(r"operating margin|^ebit margin|^op(?:erating)?\.?\s*margin", re.I), "OperatingMargin"),
    (re.compile(r"^ebitda\b(?! margin)", re.I), "EBITDA"),
    (re.compile(r"^ebitda margin", re.I), "EBITDAMargin"),
    (re.compile(r"^net income|^net profit|^earnings\b", re.I), "NetIncome"),
    (re.compile(r"net (profit )?margin", re.I), "NetMargin"),
    (re.compile(r"diluted eps|^eps\b|earnings per share", re.I), "DilutedEPS"),
    (re.compile(r"diluted shares|shares outstanding|share count", re.I), "DilutedShares"),
    (re.compile(r"free cash flow|^fcf\b", re.I), "FreeCashFlow"),
    (re.compile(r"operating cash flow|^ocf\b|cash from ops", re.I), "OperatingCashFlow"),
    (re.compile(r"^capex|capital expend", re.I), "CapEx"),
    (re.compile(r"cash (&|and) (cash )?equiv|cash & st|cash\b", re.I), "Cash"),
    (re.compile(r"total debt|^debt\b", re.I), "TotalDebt"),
    (re.compile(r"net debt", re.I), "NetDebt"),
    (re.compile(r"total assets", re.I), "TotalAssets"),
    (re.compile(r"stockholders.? equity|shareholders.? equity|^equity\b", re.I), "StockholdersEquity"),
    (re.compile(r"^dividends\b", re.I), "Dividends"),
    (re.compile(r"buyback|share repo", re.I), "BuybacksCash"),
]


def _metric_from_label(label: str) -> Optional[str]:
    s = _strip_md(label)
    s = re.sub(r"\s*\(\$?b(?:n|illions?)?\)\s*$", "", s, flags=re.I)
    s = re.sub(r"\s*\(\$?m(?:illions)?\)\s*$", "", s, flags=re.I)
    s = re.sub(r"\s*\(\$\)\s*$", "", s)
    for pat, key in _METRIC_PATTERNS:
        if pat.search(s):
            return key
    return None


def _as_rate(v: Optional[float]) -> Optional[float]:
    """13.3 or 13.3% → 0.133. Already-fractional rates pass through."""
    if v is None:
        return None
    if abs(v) > 1.5:
        return v / 100.0
    return v


def reconcile_money_level(est: Optional[float], actual: Optional[float]) -> Optional[float]:
    """Put a report estimate into $ millions next to a known actual.

    Drops growth-looking leftovers (13.3 vs 45,183) and lifts billions
    written as 47.5 next to an actual of 45,183.
    """
    if est is None:
        return None
    if actual is None or actual == 0:
        return est
    ratio = abs(est / actual)
    if 0.40 <= ratio <= 2.5:
        return est
    if 0.40 <= abs(est * 1000.0 / actual) <= 2.5:
        return est * 1000.0
    if abs(est) >= 10_000_000 and 0.40 <= abs(est / 1_000_000.0 / actual) <= 2.5:
        return est / 1_000_000.0
    if abs(est) < 80 and ratio < 0.05:
        return None
    return est


def normalize_shares_millions(
    report: Optional[float],
    fallback: Optional[float] = None,
) -> Optional[float]:
    """Diluted shares in millions. Prefer SEC; reject 3.0-as-billions leftovers."""
    v = _f(report)
    fb = _f(fallback)
    if v is not None and abs(v) >= 100_000:
        v = v / 1_000_000.0
    if fb is not None and abs(fb) >= 100_000:
        fb = fb / 1_000_000.0
    if v is not None and fb is not None and abs(fb) >= 50:
        if abs(v) < 30:
            if abs(v * 1000.0 - fb) / fb <= 0.35:
                v = v * 1000.0
            else:
                v = fb
        elif abs(v / fb) > 8 or abs(v / fb) < 0.15:
            v = fb
    if v is None:
        v = fb
    return v


def _scaled_table_number(raw: str, label: str, header: str, metric: str) -> Optional[float]:
    val = parse_cell_number(raw)
    if val is None:
        val = parse_embedded_number(raw)
    if val is None:
        return None
    if metric in ("RevenueGrowth", "FCFGrowth", "EPSGrowth") or metric in _MARGINS:
        return _as_rate(val) if metric.endswith("Growth") else val
    if metric in _PERSHARE:
        return val
    if metric in _SHARES:
        return val
    if metric in _MONEY and not _cell_has_explicit_unit(raw):
        scale = _money_scale_from_text(label) or _money_scale_from_text(header)
        if scale and scale != 1.0:
            val = val * scale
    return val


def extract_forecasts(tables: list[dict]) -> dict[int, dict[str, float]]:
    """Map fiscal year → metric → value for estimate columns only."""
    out: dict[int, dict[str, float]] = {}
    for tbl in tables:
        headers = tbl.get("headers") or []
        for col_i, h in enumerate(headers):
            if col_i == 0:
                continue
            if not _col_is_estimate(h):
                continue
            year = _year_from_header(h)
            if year is None:
                continue
            bucket = out.setdefault(year, {})
            for row in tbl.get("rows") or []:
                if not row:
                    continue
                label = row[0] if row else ""
                key = _metric_from_label(label)
                if not key or col_i >= len(row):
                    continue
                val = _scaled_table_number(row[col_i], label, h, key)
                if val is None:
                    continue
                bucket[key] = val
    return out


def extract_actuals(tables: list[dict]) -> dict[int, dict[str, float]]:
    """Same as extract_forecasts, but FY actual columns (not E / TTM)."""
    out: dict[int, dict[str, float]] = {}
    for tbl in tables:
        headers = tbl.get("headers") or []
        for col_i, h in enumerate(headers):
            if col_i == 0:
                continue
            if _col_is_estimate(h) or _col_is_ttm(h):
                continue
            year = _year_from_header(h)
            if year is None:
                continue
            bucket = out.setdefault(year, {})
            for row in tbl.get("rows") or []:
                if not row:
                    continue
                label = row[0] if row else ""
                key = _metric_from_label(label)
                if not key or col_i >= len(row):
                    continue
                val = _scaled_table_number(row[col_i], label, h, key)
                if val is None:
                    continue
                bucket[key] = val
    return out


def _cagr(start: Optional[float], end: Optional[float], years: int) -> Optional[float]:
    if start is None or end is None or not years or start <= 0:
        return None
    try:
        return (end / start) ** (1.0 / years) - 1.0
    except (ValueError, ZeroDivisionError, OverflowError):
        return None


def forward_growth(tables: list[dict]) -> dict[str, Optional[float]]:
    """Near-term and multi-year forward revenue / EPS growth from report tables."""
    act = extract_actuals(tables)
    est = extract_forecasts(tables)
    rev_g: Optional[float] = None
    eps_g: Optional[float] = None

    def _pick(cur: Optional[float], nxt: Optional[float]) -> Optional[float]:
        if nxt is None:
            return cur
        if cur is None:
            return nxt
        # Prefer the more conservative (lower) of two positive rates; if signs
        # differ, keep the near-term figure.
        if cur >= 0 and nxt >= 0:
            return max(cur, nxt)
        return nxt if abs(nxt) > abs(cur) else cur

    if act and est:
        ay, ey = max(act), min(est)
        n = max(1, ey - ay)
        rev_g = _pick(rev_g, _cagr(act[ay].get("Revenue"), est[ey].get("Revenue"), n))
        eps_g = _pick(eps_g, _cagr(act[ay].get("DilutedEPS"), est[ey].get("DilutedEPS"), n))
    years = sorted(est.keys())
    if len(years) >= 2:
        y0, y1 = years[0], years[1]
        r0, r1 = est[y0].get("Revenue"), est[y1].get("Revenue")
        if r0 and r1 and r0 > 0:
            rev_g = _pick(rev_g, r1 / r0 - 1.0)
        e0, e1 = est[y0].get("DilutedEPS"), est[y1].get("DilutedEPS")
        if e0 and e1 and e0 > 0:
            eps_g = _pick(eps_g, e1 / e0 - 1.0)
        last_y = years[-1]
        span = last_y - y0
        if span >= 2:
            rev_g = _pick(rev_g, _cagr(est[y0].get("Revenue"), est[last_y].get("Revenue"), span))
            eps_g = _pick(eps_g, _cagr(est[y0].get("DilutedEPS"), est[last_y].get("DilutedEPS"), span))
    return {"rev": rev_g, "eps": eps_g}


VALUE_CUT = 0.05    # |DCF/last − 1| ≥ 5% → cheap / rich
_FAIR_BAND = 0.05   # |value/last − 1| inside this band → FAIR VALUED
GROWTH_CUT = 0.15   # 15%+ forward rev or EPS → growth

STYLE_LABELS = {
    "value": "VALUE",
    "growth": "GROWTH",
    "garp": "GARP",
    "expensive": "RICH",
    "core": "CORE",
}


def style_from_metrics(
    dcf_value: Optional[float],
    last: Optional[float],
    rev_g: Optional[float] = None,
    eps_g: Optional[float] = None,
) -> dict[str, Any]:
    """Desk style: VALUE / GROWTH / GARP / RICH / CORE from DCF vs last + fwd growth."""
    dcf_gap = None
    if dcf_value and last and last > 0:
        dcf_gap = dcf_value / float(last) - 1.0
    undervalued = dcf_gap is not None and dcf_gap >= VALUE_CUT
    overvalued = dcf_gap is not None and dcf_gap <= -VALUE_CUT
    growing = (rev_g is not None and rev_g >= GROWTH_CUT) or (
        eps_g is not None and eps_g >= GROWTH_CUT
    )
    if dcf_gap is None and not growing:
        style = None
    elif undervalued and growing:
        style = "garp"
    elif undervalued:
        style = "value"
    elif growing:
        style = "growth"
    elif overvalued:
        style = "expensive"
    else:
        style = "core"

    bits: list[str] = []
    if dcf_gap is not None:
        bits.append(f"DCF {dcf_gap * 100:+.0f}% vs last")
    elif dcf_value is None:
        bits.append("No DCF $/sh in report")
    if rev_g is not None:
        bits.append(f"fwd rev {rev_g * 100:+.0f}%")
    if eps_g is not None:
        bits.append(f"fwd EPS {eps_g * 100:+.0f}%")
    explain = {
        "value": "DCF undervalued — cheap on the model",
        "growth": "Forward revenue/earnings growth; not yet cheap on DCF",
        "garp": "DCF cheap and growing — growth at a reasonable price",
        "expensive": "DCF above last — rich vs the model",
        "core": "Fair on DCF, no standout forward growth",
    }
    note = (explain.get(style) or "Not enough DCF / growth to classify")
    if bits:
        note = note + " · " + " · ".join(bits)
    return {
        "style": style,
        "label": STYLE_LABELS.get(style or "", ""),
        "note": note,
        "dcf_value": dcf_value,
        "dcf_gap": dcf_gap,
        "fwd_rev_growth": rev_g,
        "fwd_eps_growth": eps_g,
    }


def rank_dcf_undervalued(
    items: list[dict[str, Any]],
    *,
    limit: int = 10,
) -> list[dict[str, Any]]:
    """Top N names cheapest on DCF vs last price. DCF only — no PT / style mix.

    Each item: {ticker, dcf_value, price}. Drops missing/non-positive prices
    and anything not strictly undervalued (dcf > last). Sorted by gap desc.
    """
    ranked: list[dict[str, Any]] = []
    for it in items or []:
        tk = str((it or {}).get("ticker") or "").strip().upper()
        dcf = _f((it or {}).get("dcf_value"))
        px = _f((it or {}).get("price"))
        if not tk or dcf is None or px is None or dcf <= 0 or px <= 0:
            continue
        gap = dcf / px - 1.0
        if gap <= 0:
            continue
        ranked.append({
            "ticker": tk,
            "dcf_value": dcf,
            "price": px,
            "dcf_gap": gap,
            "dcf_gap_pct": round(gap * 100.0, 2),
        })
    ranked.sort(key=lambda r: (-r["dcf_gap"], r["ticker"]))
    return ranked[: max(0, int(limit or 0))]


def classify_stock_style(
    md: str,
    *,
    summary: Optional[dict] = None,
    price: Optional[float] = None,
) -> dict[str, Any]:
    """Parse a saved report and return style_from_metrics()."""
    summary = summary if isinstance(summary, dict) else {}
    tables = parse_md_tables(md or "")
    dcf = extract_dcf(md or "")
    wacc = extract_wacc_build(tables)
    implied = _f(dcf.get("implied_price")) or _f(wacc.get("implied_price"))
    last = _f(price) or _f(summary.get("current_price"))
    g = forward_growth(tables)
    return style_from_metrics(implied, last, g.get("rev"), g.get("eps"))


def parse_embedded_number(raw: str) -> Optional[float]:
    """Like parse_cell_number but finds the first number anywhere in the cell."""
    v = parse_cell_number(raw)
    if v is not None:
        return v
    s = _strip_md(raw)
    if not s:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", s)
    if m:
        try:
            return float(m.group(1)) / 100.0
        except ValueError:
            return None
    m = re.search(
        r"\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?|[0-9]*\.[0-9]+|[0-9]+)"
        r"\s*(bn|billions?|mm|millions?|b|m|k|thousands?|x|×)?",
        s,
        re.I,
    )
    if not m:
        return None
    try:
        val = float(m.group(1).replace(",", ""))
    except ValueError:
        return None
    unit = (m.group(2) or "").lower()
    if unit in ("x", "×"):
        pass
    else:
        val = _apply_money_unit(val, unit)
    if s.strip().startswith("(") or re.search(r"(^|\s)[-−]", s):
        val = -abs(val)
    return val


def extract_dcf(md: str) -> dict[str, Any]:
    """Pull DCF scalars from prose (tables are handled separately)."""
    text = md or ""
    out: dict[str, Any] = {}

    def _pct(pat: str) -> Optional[float]:
        m = re.search(pat, text, re.I)
        if not m:
            return None
        try:
            return float(m.group(1)) / 100.0
        except ValueError:
            return None

    def _num(pat: str) -> Optional[float]:
        m = re.search(pat, text, re.I)
        if not m:
            return None
        try:
            return float(m.group(1).replace(",", ""))
        except ValueError:
            return None

    def _money(pat: str) -> Optional[float]:
        m = re.search(pat, text, re.I)
        if not m:
            return None
        try:
            val = float(m.group(1).replace(",", ""))
        except ValueError:
            return None
        unit = (m.group(2) or "").lower() if m.lastindex and m.lastindex >= 2 else ""
        if unit in ("bn", "b"):
            val *= 1000.0
        return val

    out["rf"] = _pct(r"risk[-\s]*free(?:\s+rate)?[^%]{0,40}?(\d+(?:\.\d+)?)\s*%")
    out["erp"] = _pct(
        r"(?:equity risk premium|\bERP\b|market risk premium)[^%]{0,40}?(\d+(?:\.\d+)?)\s*%"
    )
    out["beta"] = _num(r"\bbeta\b(?:\s*\(levered\))?[^0-9]{0,24}(\d+(?:\.\d+)?)")
    out["ke"] = _pct(r"cost of equity[^%]{0,40}?(\d+(?:\.\d+)?)\s*%")
    out["kd_pretax"] = _pct(
        r"(?:pre[-\s]*tax\s+)?cost of debt[^%]{0,40}?(\d+(?:\.\d+)?)\s*%"
    )
    out["we"] = _pct(r"E\s*/\s*\(?D\s*\+\s*E\)?[^%]{0,24}?(\d+(?:\.\d+)?)\s*%")
    out["wd"] = _pct(r"D\s*/\s*\(?D\s*\+\s*E\)?[^%]{0,24}?(\d+(?:\.\d+)?)\s*%")
    out["wacc"] = _pct(r"\bWACC\b[^%]{0,48}?(\d+(?:\.\d+)?)\s*%")
    out["terminal_growth"] = _pct(
        r"terminal\s+growth(?:\s+rate)?[^%]{0,48}?(\d+(?:\.\d+)?)\s*%"
    )
    out["tax_rate"] = _pct(r"(?:tax\s+rate|effective\s+tax)[^%]{0,40}?(\d+(?:\.\d+)?)\s*%")
    sh_m = re.search(
        r"(?:diluted\s+)?shares(?:\s+outstanding)?[^0-9]{0,36}?"
        r"(\d{1,6}(?:\.\d+)?)\s*(bn|billions?|mm|millions?|b|m)?",
        text,
        re.I,
    )
    if sh_m:
        try:
            sh_val = float(sh_m.group(1).replace(",", ""))
            sh_val = _apply_money_unit(sh_val, sh_m.group(2) or "m")
            out["shares"] = sh_val
        except ValueError:
            pass
    out["net_debt"] = _money(
        r"net\s+debt[^\$]{0,40}\$\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?|[0-9]+(?:\.[0-9]+)?)"
        r"\s*(bn|b|mm|m)?"
    )
    out["year0_fcf"] = _money(
        r"(?:year[-\s]*0|base)\s+FCF[^\$]{0,40}\$\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?|[0-9]+(?:\.[0-9]+)?)"
        r"\s*(bn|b|mm|m)?"
    )
    out["enterprise_value"] = _money(
        r"(?:implied\s+)?enterprise\s+value[^\$]{0,40}\$\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?|[0-9]+(?:\.[0-9]+)?)"
        r"\s*(bn|b|mm|m)?"
    )
    out["equity_value"] = _money(
        r"(?:implied\s+)?equity\s+value[^\$]{0,40}\$\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?|[0-9]+(?:\.[0-9]+)?)"
        r"\s*(bn|b|mm|m)?"
    )
    out["terminal_value"] = _money(
        r"terminal\s+value(?:\s*\(exit\))?[^\$]{0,40}\$\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?|[0-9]+(?:\.[0-9]+)?)"
        r"\s*(bn|b|mm|m)?"
    )
    m = re.search(
        r"(?:implied|dcf)\s+(?:share\s+)?(?:price|value)[^\$]{0,40}\$\s*([0-9]{1,5}(?:,[0-9]{3})*(?:\.[0-9]+)?)",
        text,
        re.I,
    )
    if m:
        try:
            out["implied_price"] = float(m.group(1).replace(",", ""))
        except ValueError:
            pass
    m = re.search(r"DCF value\s*/\s*share[^\$]{0,20}\$\s*([0-9]{1,5}(?:,[0-9]{3})*(?:\.[0-9]+)?)", text, re.I)
    if m and out.get("implied_price") is None:
        try:
            out["implied_price"] = float(m.group(1).replace(",", ""))
        except ValueError:
            pass
    return {k: v for k, v in out.items() if v is not None}


def _fmt_dcf_pt(v: float) -> str:
    if abs(v - round(v)) < 0.049:
        return f"${round(v):,.0f}"
    return f"${v:,.2f}"


def resolve_cover_dcf(md: str, dcf_value: Optional[float] = None) -> Optional[float]:
    """DCF $/share for the cover Rating row — never the blended 12m PT.

    Prefers the model equity-bridge value/share, then the PT-derivation
    DCF (base) implied, then prose ``implied_price``.
    """
    v = _f(dcf_value)
    if v is not None and v > 1:
        return v
    tables = parse_md_tables(md or "")
    dcf = extract_dcf(md or "")
    v = extract_dcf_model_share(tables, dcf)
    if v is not None and v > 1:
        return v
    for ap in extract_valuation_approaches(md or "", tables=tables, dcf=dcf):
        if ap.get("id") in ("dcf", "dcf_base"):
            x = _f(ap.get("value"))
            if x is not None and x > 1:
                return x
    v = _f((dcf or {}).get("implied_price"))
    if v is not None and v > 1:
        return v
    return None


def inject_cover_dcf_target(md: str, dcf_value: Optional[float] = None) -> str:
    """Put the DCF-only $/share in the cover table Rating row, right cell.

    The 12-month PT row stays the blended target. Empty / conviction-pick
    placeholders are replaced. Existing DCF Target cells are left alone.
    """
    if not md:
        return md
    val = resolve_cover_dcf(md, dcf_value)
    if val is None or val <= 0:
        return md
    label = f"**DCF Target:** {_fmt_dcf_pt(val)}"
    lines = md.replace("\r\n", "\n").split("\n")
    i = 0
    n = len(lines)
    while i < n and "|" not in lines[i]:
        i += 1
    if i >= n:
        return md
    j = i
    while j < n and "|" in lines[j]:
        j += 1
    changed = False
    for k in range(i, j):
        raw = lines[k]
        if not re.search(r"rating\s*:", raw, re.I):
            continue
        if re.search(r"dcf\s+target", raw, re.I):
            break
        s = raw.rstrip("\n")
        trail_pipe = s.rstrip().endswith("|")
        body = s.strip()
        if body.startswith("|"):
            body = body[1:]
        if body.endswith("|"):
            body = body[:-1]
        cells = body.split("|")
        if len(cells) < 2:
            break
        right = _strip_md(cells[1])
        placeholder = (not right) or bool(
            re.fullmatch(r"★?\s*CONVICTION PICK", right, re.I)
        )
        if not placeholder:
            break
        cells[1] = f" {label} "
        rebuilt = "|" + "|".join(cells) + ("|" if trail_pipe else "")
        lines[k] = rebuilt
        changed = True
        break
    return "\n".join(lines) if changed else md


_WACC_LABELS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"risk[-\s]*free", re.I), "rf"),
    (re.compile(r"equity risk premium|\berp\b|market risk premium", re.I), "erp"),
    (re.compile(r"\bbeta\b", re.I), "beta"),
    (re.compile(r"cost of equity|^ke\b", re.I), "ke"),
    (re.compile(r"after[-\s]*tax cost of debt", re.I), "kd_aftertax"),
    (re.compile(r"(?:pre[-\s]*tax\s+)?cost of debt|^kd\b", re.I), "kd_pretax"),
    (re.compile(r"tax rate|effective tax", re.I), "tax_rate"),
    (re.compile(r"e\s*/\s*\(?d\s*\+\s*e|equity weight", re.I), "we"),
    (re.compile(r"d\s*/\s*\(?d\s*\+\s*e|debt weight", re.I), "wd"),
    (re.compile(r"\bwacc\b", re.I), "wacc"),
    (re.compile(r"terminal growth|^g\b", re.I), "terminal_growth"),
    (re.compile(r"net debt|net cash", re.I), "net_debt"),
    (re.compile(r"diluted shares|share count|shares \(m", re.I), "shares"),
    (re.compile(r"dcf value|implied (?:share )?price|value / share", re.I), "implied_price"),
    (re.compile(r"enterprise value", re.I), "enterprise_value"),
    (re.compile(r"equity value", re.I), "equity_value"),
    (re.compile(r"terminal value", re.I), "terminal_value"),
]


def _wacc_key(label: str) -> Optional[str]:
    s = _strip_md(label)
    for pat, key in _WACC_LABELS:
        if pat.search(s):
            return key
    return None


def extract_wacc_build(tables: list[dict]) -> dict[str, Any]:
    """Component / Value WACC build table + any KV valuation table."""
    out: dict[str, Any] = {}
    for tbl in tables:
        headers = [h.lower() for h in (tbl.get("headers") or [])]
        title = (tbl.get("title") or "").lower()
        rows = tbl.get("rows") or []
        if "sensitivity" in title or "sensitivity" in (headers[0] if headers else ""):
            continue
        if "ladder" in title or "comparable" in title or "peer" in title:
            continue
        labs = " ".join(_strip_md(r[0]) if r else "" for r in rows).lower()
        looks = (
            "wacc" in title
            or "wacc" in " ".join(headers)
            or "risk-free" in labs
            or "cost of equity" in labs
            or "component" in (headers[0] if headers else "")
            or "step" in (headers[0] if headers else "")
        )
        if not looks:
            continue
        for row in rows:
            if not row:
                continue
            key = _wacc_key(row[0])
            if not key:
                continue
            val = parse_embedded_number(row[1]) if len(row) > 1 else None
            if val is None:
                continue
            # Shares stored as millions; a raw 1.5e9 would be unusual in this table.
            if key == "beta" and val > 10:
                continue
            if key in ("we", "wd", "wacc", "rf", "erp", "ke", "kd_pretax", "kd_aftertax",
                       "tax_rate", "terminal_growth") and val > 1.5:
                val = val / 100.0
            out[key] = val
            if len(row) > 2 and _strip_md(row[2]):
                out[f"{key}_notes"] = _strip_md(row[2])
    return out


def extract_dcf_ladder(tables: list[dict]) -> list[dict[str, Any]]:
    """Year-by-year DCF Projection Ladder rows (skip totals)."""
    for tbl in tables:
        headers = tbl.get("headers") or []
        hlow = [h.lower() for h in headers]
        title = (tbl.get("title") or "").lower()
        if not (
            "ladder" in title
            or (any("fcf" in h for h in hlow) and any("discount" in h for h in hlow))
            or any("pv of fcf" in h or "pv fcf" in h for h in hlow)
        ):
            continue
        idx: dict[str, int] = {}
        for i, h in enumerate(hlow):
            if "fiscal" in h:
                idx["fy"] = i
            elif re.search(r"^year\b", h) or h.strip() == "t":
                idx["year"] = i
            elif "rev growth" in h or "revenue growth" in h:
                idx["rev_g"] = i
            elif "revenue" in h:
                idx["revenue"] = i
            elif "fcf growth" in h:
                idx["fcf_g"] = i
            elif "discount" in h:
                idx["df"] = i
            elif "pv" in h:
                idx["pv"] = i
            elif "fcf" in h:
                idx["fcf"] = i
        rows_out: list[dict[str, Any]] = []
        for row in tbl.get("rows") or []:
            if not row:
                continue
            lab = _strip_md(row[0]).lower()
            if any(k in lab for k in ("sum", "total", "pv explicit")):
                continue
            item: dict[str, Any] = {}
            year_raw = row[idx["year"]] if "year" in idx and idx["year"] < len(row) else row[0]
            ym = re.search(r"(\d+)", year_raw or "")
            if ym:
                item["t"] = int(ym.group(1))
            if "(a)" in lab or re.search(r"\bA\b", year_raw or ""):
                item["kind"] = "A"
            elif "(e)" in lab or re.search(r"\bE\b", year_raw or ""):
                item["kind"] = "E"
            if "fy" in idx and idx["fy"] < len(row):
                item["fy"] = _year_from_header(row[idx["fy"]])
            for key in ("revenue", "fcf", "rev_g", "fcf_g", "df", "pv"):
                if key not in idx or idx[key] >= len(row):
                    continue
                raw = row[idx[key]]
                val = parse_embedded_number(raw)
                if val is None:
                    continue
                if key in ("revenue", "fcf", "pv") and not _cell_has_explicit_unit(raw):
                    scale = _money_scale_from_text(headers[idx[key]] if idx[key] < len(headers) else "")
                    if scale and scale != 1.0:
                        val *= scale
                if key in ("rev_g", "fcf_g"):
                    val = _as_rate(val) if abs(val) > 1.5 else val
                item[key] = val
            if item.get("t") is None and not item.get("fcf") and not item.get("revenue"):
                continue
            rows_out.append(item)
        if rows_out:
            return rows_out
    return []


def extract_sensitivity(tables: list[dict]) -> Optional[dict[str, Any]]:
    """WACC × terminal-growth sensitivity grid of DCF $/share."""
    for tbl in tables:
        headers = tbl.get("headers") or []
        title = (tbl.get("title") or "").lower()
        h0 = (headers[0] if headers else "").lower()
        if "sensitivity" not in title and "sensitivity" not in h0:
            if not any(re.search(r"\btgr\b|terminal", h, re.I) for h in headers[1:]):
                continue
        tgrs: list[float] = []
        for h in headers[1:]:
            v = parse_embedded_number(h)
            if v is None:
                continue
            if v > 1:
                v = v / 100.0
            tgrs.append(v)
        if len(tgrs) < 2:
            continue
        waccs: list[float] = []
        cells: dict[tuple[float, float], float] = {}
        for row in tbl.get("rows") or []:
            if not row:
                continue
            w = parse_embedded_number(row[0])
            if w is None:
                continue
            if w > 1:
                w = w / 100.0
            waccs.append(w)
            for i, g in enumerate(tgrs):
                if i + 1 >= len(row):
                    continue
                px = parse_embedded_number(row[i + 1])
                if px is None:
                    continue
                cells[(round(w, 4), round(g, 4))] = px
        if waccs and cells:
            return {"waccs": waccs, "tgrs": tgrs, "cells": cells, "table": tbl}
    return None


def extract_bridge_table(tables: list[dict]) -> Optional[dict]:
    for tbl in tables:
        title = (tbl.get("title") or "").lower()
        headers = " ".join(h.lower() for h in (tbl.get("headers") or []))
        labs = " ".join(_strip_md(r[0]) if r else "" for r in (tbl.get("rows") or [])).lower()
        if "bridge" in title or (
            "terminal value" in labs and "equity" in labs
        ) or ("step" in headers and "amount" in headers):
            return tbl
    return None


def extract_scenarios(md: str) -> list[dict[str, Any]]:
    """Bull / base / bear price targets + probabilities from the verdict."""
    text = md or ""
    rows = []
    for label, pat in (
        ("Bull", r"bull\s+case[^\$\n]{0,90}\$\s*([0-9]{1,5}(?:,[0-9]{3})*(?:\.[0-9]+)?)"),
        ("Base", r"base\s+case[^\$\n]{0,90}\$\s*([0-9]{1,5}(?:,[0-9]{3})*(?:\.[0-9]+)?)"),
        ("Bear", r"bear\s+case[^\$\n]{0,90}\$\s*([0-9]{1,5}(?:,[0-9]{3})*(?:\.[0-9]+)?)"),
    ):
        m = re.search(pat, text, re.I)
        if not m:
            continue
        try:
            pt = float(m.group(1).replace(",", ""))
        except ValueError:
            continue
        window = text[max(0, m.start() - 40) : m.end() + 80]
        pm = re.search(r"(\d+(?:\.\d+)?)\s*%", window)
        prob = None
        if pm:
            try:
                prob = float(pm.group(1)) / 100.0
            except ValueError:
                prob = None
        rows.append({"case": label, "price_target": pt, "probability": prob})
    return rows


def extract_street_table(tables: list[dict]) -> Optional[dict]:
    for tbl in tables:
        headers = [h.lower() for h in (tbl.get("headers") or [])]
        joined = " ".join(headers)
        if "firm" in joined and ("target" in joined or "rating" in joined):
            return tbl
        title = (tbl.get("title") or "").lower()
        if "analyst" in title or "street" in title or "consensus" in title:
            if tbl.get("rows"):
                return tbl
    return None


def extract_comps_table(tables: list[dict]) -> Optional[dict]:
    for tbl in tables:
        headers = [h.lower() for h in (tbl.get("headers") or [])]
        joined = " ".join(headers)
        if any(k in joined for k in ("ev/ebitda", "p/e", "pe ", "fcf yield", "ev/sales")):
            title = (tbl.get("title") or "").lower()
            if "peer" in title or "comp" in title or "comparable" in joined or "ticker" in joined:
                return tbl
            if "peer" in joined or "company" in joined or "ticker" in joined:
                return tbl
    return None


def extract_derivation_table(tables: list[dict]) -> Optional[dict]:
    for tbl in tables:
        title = (tbl.get("title") or "").lower()
        headers = " ".join(h.lower() for h in (tbl.get("headers") or []))
        if "derivation" in title or "price target" in title and "method" in headers:
            return tbl
        if "method" in headers and "weight" in headers:
            return tbl
    return None


def valuation_verdict(
    value: Optional[float],
    last: Optional[float],
    *,
    band: float = _FAIR_BAND,
) -> dict[str, Any]:
    """UNDERVALUED / FAIR VALUED / OVERVALUED vs last. No blending.

    intensity 0..1 scales how far the name is from fair (for color gradients).
    Positive gap = cheap (value > last) = undervalued.
    """
    v = _f(value)
    px = _f(last)
    if v is None or px is None or v <= 0 or px <= 0:
        return {
            "verdict": None, "label": "—", "tone": "none",
            "gap": None, "intensity": 0.0, "value": v, "last": px,
        }
    gap = v / px - 1.0
    ag = abs(gap)
    if ag < band:
        tone, label = "fair", "FAIR VALUED"
        intensity = 0.0
    elif gap > 0:
        tone, label = "under", "UNDERVALUED"
        intensity = min(1.0, max(0.2, (ag - band) / 0.35))
    else:
        tone, label = "over", "OVERVALUED"
        intensity = min(1.0, max(0.2, (ag - band) / 0.35))
    return {
        "verdict": label,
        "label": label,
        "tone": tone,
        "gap": gap,
        "intensity": round(float(intensity), 3),
        "value": v,
        "last": px,
    }


def verdict_palette(tone: Optional[str], intensity: Optional[float] = 0.0) -> tuple[str, str]:
    """(fill_hex, font_hex). Dark text on light fills — never white-on-white."""
    i = 0.0 if intensity is None else max(0.0, min(1.0, float(intensity)))
    t = (tone or "").lower()
    if t in ("under", "undervalued"):
        if i < 0.35:
            return "D1FAE5", "14532D"
        if i < 0.60:
            return "6EE7B7", "064E3B"
        if i < 0.80:
            return "059669", "ECFDF5"
        return "047857", "ECFDF5"
    if t in ("over", "overvalued"):
        if i < 0.35:
            return "FEE2E2", "7F1D1D"
        if i < 0.60:
            return "FCA5A5", "7F1D1D"
        if i < 0.80:
            return "DC2626", "FEF2F2"
        return "B91C1C", "FEF2F2"
    if t in ("fair", "fair valued"):
        return "E2E8F0", "0F172A"
    return "F1F5F9", "0F172A"


def _approach_id(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "", (name or "").lower())
    aliases = {
        "dcfvalueshare": "dcf", "dcfvalue": "dcf",
        "dcfbase": "dcf_base", "basedcfd": "dcf_base", "basedcf": "dcf_base",
        "comparable": "comps", "comparables": "comps", "tradingcomps": "comps",
        "peermultiples": "comps", "multiples": "comps",
        "pecomps": "pe", "pe": "pe",
        "evrevenue": "ev_rev", "evrevenucomps": "ev_rev", "evsales": "ev_rev",
        "streetanchored": "street_anchored", "streetanchor": "street_anchored",
        "streetconsensus": "street", "consensus": "street", "analystpt": "street",
        "bullcase": "bull", "basecase": "base", "bearcase": "bear",
        "report12mpt": "report_pt", "12monthpricetarget": "report_pt",
        "dgatarget": "report_pt",
    }
    return aliases.get(s, s or "other")


def _norm_method_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def _is_dcf_base_method(name: str) -> bool:
    n = _norm_method_name(name)
    return bool(re.fullmatch(r"dcf(\s*\(\s*base\s*\))?|dcf base|base[- ]case dcf", n))


def _is_extra_dcf_method(name: str) -> bool:
    """Skip duration / bull-case DCF variants — keep model share + DCF (base) only."""
    n = _norm_method_name(name)
    if "dcf" not in n and "discounted cash" not in n:
        return False
    if _is_dcf_base_method(name):
        return False
    if "value / share" in n or "value/share" in n:
        return False
    return True


def _is_blend_pt_row(name: str) -> bool:
    n = _norm_method_name(name)
    return bool(re.search(
        r"12-?m(?:onth)?\s+price\s+target|blend|weighted\s+avg|weighted average|"
        r"total pt|dga target|expected value",
        n,
    ))


def _derivation_columns(headers: list[str]) -> dict[str, int]:
    """Prefer Implied Value over Weighted Value. Old 3-col tables use Value."""
    out: dict[str, int] = {}
    for i, h in enumerate(headers or []):
        hl = (h or "").lower()
        if "implied" in hl:
            out["implied"] = i
        elif "weighted" in hl and ("value" in hl or "price" in hl or "$" in hl):
            out["weighted"] = i
        elif re.search(r"\bweight", hl) and "value" not in hl and "price" not in hl:
            out["weight"] = i
        elif "value" in hl or hl in ("price", "$/sh", "$ / share", "pt"):
            out.setdefault("value", i)
    return out


def extract_dcf_model_share(
    tables: list[dict],
    dcf: Optional[dict] = None,
) -> Optional[float]:
    """Model DCF value / share from the equity bridge — not PT-derivation DCF (base)."""
    br = extract_bridge_table(tables)
    if br:
        for row in br.get("rows") or []:
            if not row:
                continue
            lab = _strip_md(row[0]).lower()
            if "dcf value" in lab and "share" in lab:
                v = parse_embedded_number(row[1]) if len(row) > 1 else None
                if v is not None and v > 1:
                    return v
    wacc = extract_wacc_build(tables)
    v = _f((wacc or {}).get("implied_price"))
    if v is not None and v > 1:
        return v
    return _f((dcf or {}).get("implied_price"))


def _street_avg_pt(tbl: Optional[dict]) -> tuple[Optional[float], int]:
    if not tbl:
        return None, 0
    headers = [h.lower() for h in (tbl.get("headers") or [])]
    idx = None
    for i, h in enumerate(headers):
        if "target" in h or re.search(r"\bpt\b", h) or "price target" in h:
            idx = i
            break
    if idx is None:
        return None, 0
    vals: list[float] = []
    for row in tbl.get("rows") or []:
        if not row or idx >= len(row):
            continue
        v = parse_embedded_number(row[idx])
        # Skip % cells and tiny numbers that are ratings, not prices.
        if v is None or v <= 1.5:
            continue
        vals.append(v)
    if not vals:
        return None, 0
    return sum(vals) / len(vals), len(vals)


def extract_valuation_approaches(
    md: str,
    *,
    last: Optional[float] = None,
    pt: Optional[float] = None,
    tables: Optional[list] = None,
    dcf: Optional[dict] = None,
) -> list[dict[str, Any]]:
    """Each method vs last on its own implied $/share. Not a weighted blend.

    DCF rows: model DCF value/share + derivation DCF (base) implied only.
    Extra DCF variants (duration / bull WACC) are dropped. EV/Revenue, P/E,
    street-anchored use Implied Value. The 12m PT is the only weighted blend.
    """
    tables = tables if tables is not None else parse_md_tables(md or "")
    dcf = dcf if dcf is not None else extract_dcf(md or "")
    last = _f(last)
    out: list[dict[str, Any]] = []
    seen: set[str] = set()

    def add(name: str, value: Optional[float], note: str = "", *, key: str = "") -> None:
        v = _f(value)
        if v is None or v <= 0:
            return
        aid = key or _approach_id(name)
        if aid in seen:
            return
        if _is_blend_pt_row(name) and aid != "report_pt":
            return
        seen.add(aid)
        vd = valuation_verdict(v, last)
        out.append({
            "id": aid,
            "name": name,
            "value": v,
            "note": note or "",
            "verdict": vd["verdict"],
            "tone": vd["tone"],
            "gap": vd["gap"],
            "intensity": vd["intensity"],
            "last": vd["last"],
        })

    model_share = extract_dcf_model_share(tables, dcf)
    add(
        "DCF value / share",
        model_share,
        "model DCF (ladder / Gordon) — not the PT-derivation DCF (base)",
        key="dcf",
    )

    deriv = extract_derivation_table(tables)
    if deriv:
        cols = _derivation_columns(deriv.get("headers") or [])
        implied_i = cols.get("implied")
        value_i = cols.get("value")
        weight_i = cols.get("weight")
        method_i = 0
        for row in deriv.get("rows") or []:
            if not row:
                continue
            method = _strip_md(row[method_i])
            if not method or re.search(r"^method$", method, re.I):
                continue
            if _is_blend_pt_row(method) or _is_extra_dcf_method(method):
                continue
            idx = implied_i if implied_i is not None else value_i
            val = parse_embedded_number(row[idx]) if idx is not None and idx < len(row) else None
            wnote = ""
            if weight_i is not None and weight_i < len(row):
                w = _strip_md(row[weight_i])
                if w and w not in ("—", "-", "–"):
                    wnote = f"report weight {w} (not applied to this verdict)"
            if _is_dcf_base_method(method):
                note = "PT derivation implied DCF (base) — not weighted; not the model DCF value / share"
                if wnote:
                    note = note + " · " + wnote
                add("DCF (base)", val, note, key="dcf_base")
                continue
            add(method, val, wnote or "implied value (not weighted)")

    street_avg, n_firms = _street_avg_pt(extract_street_table(tables))
    if n_firms:
        add("Street consensus", street_avg, f"average of {n_firms} firm PTs (implied)", key="street")

    # Bull / Base / Bear stay on the Scenarios sheet — not here. "Base case"
    # next to "DCF (base)" on Market Pulse / this table is confusing.

    rpt = _f(pt)
    add(
        "12m price target",
        rpt,
        "blend of weighted values — the only combined figure",
        key="report_pt",
    )
    return out


def recut_approaches_vs_last(
    approaches: list[dict[str, Any]],
    last: Optional[float],
) -> list[dict[str, Any]]:
    """Recompute gap / verdict / intensity vs a live last price."""
    out: list[dict[str, Any]] = []
    for raw in approaches or []:
        if not isinstance(raw, dict):
            continue
        item = dict(raw)
        vd = valuation_verdict(item.get("value"), last)
        item["verdict"] = vd["verdict"]
        item["tone"] = vd["tone"]
        item["gap"] = vd["gap"]
        item["intensity"] = vd["intensity"]
        item["last"] = vd["last"]
        out.append(item)
    return out


def approaches_from_scalars(
    *,
    dcf: Optional[float] = None,
    pt: Optional[float] = None,
    last: Optional[float] = None,
) -> list[dict[str, Any]]:
    """Minimal approaches when we only stored DCF + 12m PT (older reports)."""
    return extract_valuation_approaches(
        "",
        last=last,
        pt=pt,
        tables=[],
        dcf={"implied_price": dcf} if dcf else {},
    )


# ── Financials helpers ──────────────────────────────────────────────────────
_MONEY = {
    "Revenue", "CostOfRevenue", "GrossProfit", "OperatingIncome", "NetIncome",
    "EBITDA", "OperatingCashFlow", "CapEx", "FreeCashFlow", "Cash",
    "ShortTermInvestments", "TotalAssets", "TotalLiabilities",
    "StockholdersEquity", "LongTermDebt", "ShortTermDebt", "TotalDebt",
    "NetDebt", "Dividends", "BuybacksCash", "RnD", "DepreciationAmortization",
}
_SHARES = {"DilutedShares", "SharesOutstanding"}
_MARGINS = {"GrossMargin", "OperatingMargin", "NetMargin", "EBITDAMargin"}
_PERSHARE = {"DilutedEPS", "BasicEPS"}


def _f(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        x = float(v)
    except (TypeError, ValueError):
        return None
    if x != x:  # NaN
        return None
    return x


def to_model_units(metric: str, raw: Any, *, already_millions: bool = False) -> Optional[float]:
    """DB dollars → $ millions; margins as fractions; EPS unchanged."""
    v = _f(raw)
    if v is None:
        return None
    if metric in _MARGINS:
        return v if abs(v) <= 2 else v / 100.0
    if metric in _PERSHARE:
        return v
    if metric in _SHARES:
        return v / 1_000_000.0 if abs(v) >= 100_000 else v
    if metric in _MONEY:
        if already_millions:
            return v
        return v / 1_000_000.0
    return v


def _period_money(period: dict, metric: str, *, already_millions: bool = False) -> Optional[float]:
    if not period:
        return None
    return to_model_units(metric, period.get(metric), already_millions=already_millions)


# ── Workbook styles ─────────────────────────────────────────────────────────
def _styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color=LINE)
    med = Side(style="medium", color=NAVY)
    return {
        "navy_fill": PatternFill("solid", fgColor=NAVY),
        "gold_fill": PatternFill("solid", fgColor=GOLD),
        "section_fill": PatternFill("solid", fgColor=SECTION),
        "alt_fill": PatternFill("solid", fgColor=ROW_ALT),
        "pale_gold": PatternFill("solid", fgColor=PALE_GOLD),
        "pale_navy": PatternFill("solid", fgColor=PALE_NAVY),
        "input_fill": PatternFill("solid", fgColor=YELLOW),
        "white_fill": PatternFill("solid", fgColor=WHITE),
        "font_cover": Font(name="Calibri", size=18, bold=True, color=WHITE),
        "font_banner": Font(name="Calibri", size=10, bold=True, color=NAVY),
        "font_name": Font(name="Calibri", size=16, bold=True, color=NAVY),
        "font_h": Font(name="Calibri", size=10, bold=True, color=WHITE),
        "font_section": Font(name="Calibri", size=10, bold=True, color=WHITE),
        "font": Font(name="Calibri", size=10, color=INK),
        "font_bold": Font(name="Calibri", size=10, bold=True, color=INK),
        "font_muted": Font(name="Calibri", size=9, italic=True, color=MUTED),
        "font_est": Font(name="Calibri", size=10, color=BLUE_EST),
        "font_est_bold": Font(name="Calibri", size=10, bold=True, color=BLUE_EST),
        "font_kpi": Font(name="Calibri", size=12, bold=True, color=NAVY),
        "font_kpi_lab": Font(name="Calibri", size=8, bold=True, color=MUTED),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "center": Alignment(horizontal="center", vertical="center"),
        "right": Alignment(horizontal="right", vertical="center"),
        "thin": Border(left=thin, right=thin, top=thin, bottom=thin),
        "bottom": Border(bottom=thin),
        "top_med": Border(top=med),
        "thin_side": thin,
    }


_FMT_MM = '#,##0.0;(#,##0.0);"—"'
_FMT_SH = '$#,##0.00;($#,##0.00);"—"'
_FMT_PCT = '0.0%;(0.0%);"—"'
_FMT_X = '0.0x;(0.0x);"—"'
_FMT_SHARES = '#,##0.0;"—"'
_FMT_INT = '#,##0;(#,##0);"—"'


def _set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _print_setup(ws, *, landscape: bool = True, fit_width: int = 1) -> None:
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID if landscape else ws.PAPERSIZE_LETTER
    ws.page_setup.horizontalCentered = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddHeader.left.text = "DGA Capital Research"
    ws.oddHeader.right.text = "Confidential"
    ws.oddFooter.left.text = "Not investment advice · For intended recipient only"
    ws.oddFooter.right.text = "Page &P of &N"


def _write_banner(ws, ncols: int, title: str, subtitle: str, S) -> None:
    from openpyxl.utils import get_column_letter

    last = get_column_letter(max(ncols, 2))
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value = title
    c.font = S["font_cover"]
    c.fill = S["navy_fill"]
    c.alignment = S["left"]
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A2:{last}2")
    c = ws["A2"]
    c.value = subtitle
    c.font = S["font_banner"]
    c.fill = S["gold_fill"]
    c.alignment = S["left"]
    ws.row_dimensions[2].height = 16
    for col in range(1, max(ncols, 2) + 1):
        ws.cell(1, col).fill = S["navy_fill"]
        ws.cell(2, col).fill = S["gold_fill"]


# ── Column plan ─────────────────────────────────────────────────────────────
def _annuals_oldest_first(financials: dict) -> list[dict]:
    annuals = list((financials or {}).get("annuals") or [])
    annuals = [a for a in annuals if isinstance(a, dict)]
    annuals.sort(key=lambda a: (a.get("fy") or 0, a.get("end") or ""))
    return annuals[-8:]


def _build_columns(financials: dict, forecasts: dict[int, dict[str, float]]) -> list[dict]:
    cols: list[dict] = []
    annuals = _annuals_oldest_first(financials)
    last_fy = None
    for a in annuals:
        fy = a.get("fy")
        try:
            fy_i = int(fy) if fy is not None else None
        except (TypeError, ValueError):
            fy_i = None
        last_fy = fy_i or last_fy
        label = f"FY{fy_i}A" if fy_i else str(a.get("end") or "FY")[:7]
        cols.append({
            "label": label,
            "kind": "A",
            "fy": fy_i,
            "end": a.get("end"),
            "period": a,
            "already_mm": False,
        })
    ttm = (financials or {}).get("ttm") or {}
    if ttm and (ttm.get("Revenue") is not None or ttm.get("NetIncome") is not None):
        end = str(ttm.get("end") or "")[:7]
        cols.append({
            "label": f"TTM {end}".strip(),
            "kind": "TTM",
            "fy": None,
            "end": ttm.get("end"),
            "period": ttm,
            "already_mm": False,
        })
    # Pro forma years after the last actual FY
    last_actual = annuals[-1] if annuals else {}
    min_est_year = (last_fy + 1) if last_fy else 0
    last_rev = to_model_units("Revenue", last_actual.get("Revenue"), already_millions=False)
    for year in sorted(forecasts.keys()):
        if year < min_est_year:
            continue
        period = dict(forecasts[year] or {})
        for metric in _MONEY:
            if metric in period:
                actual = to_model_units(metric, last_actual.get(metric), already_millions=False)
                period[metric] = reconcile_money_level(period[metric], actual)
        g = _as_rate(period.get("RevenueGrowth"))
        if period.get("Revenue") is None and last_rev and last_fy and g is not None:
            span = max(1, year - last_fy)
            period["Revenue"] = last_rev * ((1.0 + g) ** span)
        cols.append({
            "label": f"FY{year}E",
            "kind": "E",
            "fy": year,
            "end": None,
            "period": period,
            "already_mm": True,
        })
    return cols


# ── Line items ──────────────────────────────────────────────────────────────
# kind: value | yoy | margin | netdebt | section
_MODEL_ROWS: list[tuple] = [
    ("section", "INCOME STATEMENT", None, None),
    ("value", "Revenue", "Revenue", "mm"),
    ("yoy", "    YoY growth", "Revenue", "pct"),
    ("value", "Cost of revenue", "CostOfRevenue", "mm"),
    ("value", "Gross profit", "GrossProfit", "mm"),
    ("margin", "    Gross margin", "GrossProfit", "Revenue"),
    ("value", "Operating income", "OperatingIncome", "mm"),
    ("margin", "    Operating margin", "OperatingIncome", "Revenue"),
    ("value", "EBITDA", "EBITDA", "mm"),
    ("margin", "    EBITDA margin", "EBITDA", "Revenue"),
    ("value", "Net income", "NetIncome", "mm"),
    ("margin", "    Net margin", "NetIncome", "Revenue"),
    ("value", "Diluted EPS", "DilutedEPS", "sh"),
    ("yoy", "    EPS growth", "DilutedEPS", "pct"),
    ("value", "Diluted shares (m)", "DilutedShares", "shares"),
    ("section", "BALANCE SHEET", None, None),
    ("value", "Cash & equivalents", "Cash", "mm"),
    ("value", "Total assets", "TotalAssets", "mm"),
    ("value", "Total debt", "TotalDebt", "mm"),
    ("netdebt", "Net debt", None, "mm"),
    ("value", "Shareholders' equity", "StockholdersEquity", "mm"),
    ("section", "CASH FLOW", None, None),
    ("value", "Operating cash flow", "OperatingCashFlow", "mm"),
    ("value", "Capital expenditure", "CapEx", "mm"),
    ("value", "Free cash flow", "FreeCashFlow", "mm"),
    ("margin", "    FCF margin", "FreeCashFlow", "Revenue"),
    ("value", "Dividends", "Dividends", "mm"),
    ("value", "Share buybacks", "BuybacksCash", "mm"),
]


def _fmt_for(kind: str) -> str:
    if kind in ("mm",):
        return _FMT_MM
    if kind == "sh":
        return _FMT_SH
    if kind == "pct":
        return _FMT_PCT
    if kind == "shares":
        return _FMT_SHARES
    return _FMT_MM


# ── Sheet builders ──────────────────────────────────────────────────────────
def _cover_sheet(wb, *, ticker, entity, sector, industry, as_of, engine,
                 rating, pt, price, upside, thesis, capital, source, dropbox_note, S):
    ws = wb.active
    ws.title = "Cover"
    _write_banner(
        ws, 8,
        "DGA CAPITAL  ·  RESEARCH MODEL",
        f"CONFIDENTIAL  ·  {ticker}  ·  {as_of}  ·  {engine or 'Research'} engine",
        S,
    )
    ws.merge_cells("A4:D4")
    ws["A4"].value = entity or ticker
    ws["A4"].font = S["font_name"]
    ws["E4"].value = ticker
    ws["E4"].font = S["font_kpi"]
    ws["E4"].alignment = S["center"]
    ws["F4"].value = (sector or "") + ((" · " + industry) if industry else "")
    ws["F4"].font = S["font_muted"]
    ws.merge_cells("F4:H4")

    kpis = [
        ("Rating", rating or "—", None),
        ("12-month PT", pt, "sh"),
        ("Last price", price, "sh"),
        ("Upside / (downside)", upside, "pct"),
        ("Market cap ($m)", capital.get("market_cap"), "mm"),
        ("Enterprise value ($m)", capital.get("enterprise_value"), "mm"),
    ]
    for i, (lab, val, fmt) in enumerate(kpis):
        col = 1 + i
        cell_l = ws.cell(6, col, lab.upper())
        cell_l.font = S["font_kpi_lab"]
        cell_l.fill = S["pale_navy"]
        cell_l.alignment = S["center"]
        cell_v = ws.cell(7, col, val if val not in ("—",) else "—")
        cell_v.font = S["font_kpi"]
        cell_v.alignment = S["center"]
        cell_v.fill = S["pale_gold"]
        cell_v.border = S["thin"]
        cell_l.border = S["thin"]
        if fmt == "sh" and isinstance(val, (int, float)):
            cell_v.number_format = _FMT_SH
        elif fmt == "pct" and isinstance(val, (int, float)):
            cell_v.number_format = _FMT_PCT
        elif fmt == "mm" and isinstance(val, (int, float)):
            cell_v.number_format = _FMT_MM
        ws.row_dimensions[7].height = 22

    ws["A9"].value = "INVESTMENT THESIS"
    ws["A9"].font = S["font_section"]
    ws["A9"].fill = S["section_fill"]
    ws.merge_cells("A9:H9")
    for col in range(1, 9):
        ws.cell(9, col).fill = S["section_fill"]
    ws.merge_cells("A10:H12")
    ws["A10"].value = thesis or "—"
    ws["A10"].alignment = Alignment_wrap(S)
    ws["A10"].font = S["font"]
    ws.row_dimensions[10].height = 36
    ws.row_dimensions[11].height = 18
    ws.row_dimensions[12].height = 18

    ws["A14"].value = "CAPITAL STRUCTURE"
    ws["A14"].font = S["font_section"]
    ws["A14"].fill = S["section_fill"]
    ws.merge_cells("A14:C14")
    for col in range(1, 4):
        ws.cell(14, col).fill = S["section_fill"]

    cap_rows = [
        ("Last price", capital.get("price"), _FMT_SH),
        ("Shares (m)", capital.get("shares"), _FMT_SHARES),
        ("Market capitalization ($m)", capital.get("market_cap"), _FMT_MM),
        ("  Cash & ST investments", capital.get("cash"), _FMT_MM),
        ("  Total debt", capital.get("total_debt"), _FMT_MM),
        ("  Net debt / (cash)", capital.get("net_debt"), _FMT_MM),
        ("Enterprise value ($m)", capital.get("enterprise_value"), _FMT_MM),
        ("Book value / sh", capital.get("book_value_ps"), _FMT_SH),
        ("P / E (FY)", capital.get("pe"), _FMT_X),
        ("EV / EBITDA (FY)", capital.get("ev_ebitda"), _FMT_X),
        ("FCF yield (FY)", capital.get("fcf_yield"), _FMT_PCT),
    ]
    r = 15
    ws.cell(r, 1, "Item").font = S["font_h"]
    ws.cell(r, 1).fill = S["navy_fill"]
    ws.cell(r, 2, "Value").font = S["font_h"]
    ws.cell(r, 2).fill = S["navy_fill"]
    ws.cell(r, 2).alignment = S["center"]
    r = 16
    for lab, val, fmt in cap_rows:
        ws.cell(r, 1, lab).font = S["font_bold"] if not lab.startswith(" ") else S["font"]
        ws.cell(r, 1).border = S["thin"]
        c = ws.cell(r, 2, val)
        c.font = S["font"]
        c.border = S["thin"]
        c.alignment = S["right"]
        if isinstance(val, (int, float)):
            c.number_format = fmt
        if r % 2 == 0:
            ws.cell(r, 1).fill = S["alt_fill"]
            c.fill = S["alt_fill"]
        r += 1

    ws["E14"].value = "MODEL NOTES"
    ws["E14"].font = S["font_section"]
    ws["E14"].fill = S["section_fill"]
    ws.merge_cells("E14:H14")
    for col in range(5, 9):
        ws.cell(14, col).fill = S["section_fill"]
    notes = [
        "Actuals (A) are SEC XBRL from the Financials store, $ millions.",
        "TTM is the stored bridge (latest FY + YTD delta) when available.",
        "Estimates (E, blue) are pro forma figures parsed from the saved report.",
        "Valuation, DCF inputs, Street targets, and scenarios come from the report.",
        "Yellow cells on Valuation are model inputs pulled from the write-up.",
        source or "Source: company_financials + saved research report.",
        dropbox_note or "",
        "DGA Capital · Confidential · Not investment advice.",
    ]
    rr = 16
    for nline in notes:
        if not nline:
            continue
        ws.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=8)
        ws.cell(rr, 5, nline).font = S["font_muted"]
        rr += 1

    _set_col_widths(ws, {ch: 22 for ch in "ABCDEFGH"})
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    _print_setup(ws, landscape=False, fit_width=1)
    ws.sheet_properties.tabColor = GOLD
    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:2"


def Alignment_wrap(S):
    from openpyxl.styles import Alignment
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


def _model_sheet(wb, cols: list[dict], S) -> dict[str, int]:
    """Write the 3-statement model. Returns {row_key: excel_row} for formulas."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Financial Model")
    n = 1 + len(cols)
    _write_banner(
        ws, max(n, 4),
        "FINANCIAL MODEL  ·  ACTUALS AND PRO FORMA",
        "$ in millions except per-share items  ·  Blue = estimate  ·  Black = reported",
        S,
    )
    # Period-end sub-header
    ws.cell(4, 1, "Period end").font = S["font_muted"]
    header_row = 5
    ws.cell(header_row, 1, "($ millions)").font = S["font_h"]
    ws.cell(header_row, 1).fill = S["navy_fill"]
    for i, col in enumerate(cols, start=2):
        cell = ws.cell(header_row, i, col["label"])
        cell.font = S["font_h"]
        cell.fill = S["gold_fill"] if col["kind"] == "E" else S["navy_fill"]
        cell.alignment = S["center"]
        cell.border = S["thin"]
        end = col.get("end")
        if end:
            e = ws.cell(4, i, str(end)[:10])
            e.font = S["font_muted"]
            e.alignment = S["center"]
    for col_i in range(1, n + 1):
        ws.cell(header_row, col_i).fill = (
            S["gold_fill"] if col_i > 1 and cols[col_i - 2]["kind"] == "E" else S["navy_fill"]
        )
        ws.cell(header_row, col_i).font = S["font_h"]
        ws.cell(header_row, col_i).border = S["thin"]

    row_index: dict[str, int] = {}
    excel_row = 6
    value_rows: dict[str, int] = {}  # metric → row for formulas

    for spec in _MODEL_ROWS:
        kind, label, metric, extra = spec
        if kind == "section":
            ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=max(n, 2))
            cell = ws.cell(excel_row, 1, label)
            cell.font = S["font_section"]
            cell.fill = S["section_fill"]
            for c in range(1, n + 1):
                ws.cell(excel_row, c).fill = S["section_fill"]
                ws.cell(excel_row, c).font = S["font_section"]
            excel_row += 1
            continue

        lab_cell = ws.cell(excel_row, 1, label)
        lab_cell.font = S["font_bold"] if kind == "value" else S["font"]
        lab_cell.alignment = S["left"]
        lab_cell.border = S["thin"]

        for i, col in enumerate(cols, start=2):
            letter = get_column_letter(i)
            is_est = col["kind"] == "E"
            font = S["font_est"] if is_est else S["font"]
            fill = S["input_fill"] if is_est and kind == "value" else None
            cell = ws.cell(excel_row, i)
            cell.font = font
            cell.alignment = S["right"]
            cell.border = S["thin"]
            if fill:
                cell.fill = fill
            elif excel_row % 2 == 0:
                cell.fill = S["alt_fill"]
                if i == 1:
                    pass

            if kind == "value":
                val = _period_money(col["period"], metric, already_millions=col.get("already_mm") or False)
                cell.value = val
                cell.number_format = _fmt_for(extra)
            elif kind == "yoy":
                src_row = value_rows.get(metric)
                if src_row and i > 2:
                    prev = get_column_letter(i - 1)
                    cell.value = (
                        f'=IF(OR({prev}{src_row}="",{prev}{src_row}=0),"—",'
                        f'({letter}{src_row}-{prev}{src_row})/{prev}{src_row})'
                    )
                else:
                    cell.value = None
                cell.number_format = _FMT_PCT
            elif kind == "margin":
                num_row = value_rows.get(metric)
                den_row = value_rows.get(extra)
                if num_row and den_row:
                    cell.value = (
                        f'=IF(OR({letter}{den_row}="",{letter}{den_row}=0),"—",'
                        f'{letter}{num_row}/{letter}{den_row})'
                    )
                else:
                    # Fall back to stored margin if we have it
                    stored = _period_money(col["period"], metric if metric.endswith("Margin") else "",
                                           already_millions=True)
                    if stored is None and metric:
                        # metric is numerator name; stored margin keys
                        mk = {
                            "GrossProfit": "GrossMargin",
                            "OperatingIncome": "OperatingMargin",
                            "EBITDA": "EBITDAMargin",
                            "NetIncome": "NetMargin",
                            "FreeCashFlow": None,
                        }.get(metric)
                        if mk:
                            stored = _period_money(col["period"], mk, already_millions=False)
                    cell.value = stored
                cell.number_format = _FMT_PCT
            elif kind == "netdebt":
                debt_r = value_rows.get("TotalDebt")
                cash_r = value_rows.get("Cash")
                if debt_r and cash_r:
                    cell.value = f'=IF(AND({letter}{debt_r}="",{letter}{cash_r}=""),"—",IF({letter}{debt_r}="",0,{letter}{debt_r})-IF({letter}{cash_r}="",0,{letter}{cash_r}))'
                else:
                    cell.value = _period_money(col["period"], "NetDebt",
                                               already_millions=col.get("already_mm") or False)
                cell.number_format = _FMT_MM

        if kind == "value" and metric:
            value_rows[metric] = excel_row
            row_index[metric] = excel_row
        excel_row += 1

    # Source line
    excel_row += 1
    ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=max(n, 2))
    ws.cell(excel_row, 1, "A = reported (SEC XBRL)   TTM = trailing twelve months   E = DGA pro forma from the saved report. Yellow cells are estimate inputs.").font = S["font_muted"]

    ws.column_dimensions["A"].width = 28
    for i in range(2, n + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.freeze_panes = "B6"
    ws.print_title_rows = "1:5"
    ws.print_title_cols = "A:A"
    _print_setup(ws, landscape=True)
    ws.sheet_properties.tabColor = NAVY
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max(n, 2))}{header_row}"
    return row_index


def _write_kv_table(ws, start_row: int, title: str, rows: list[tuple], S, ncols=4) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
    ws.cell(start_row, 1, title).font = S["font_section"]
    ws.cell(start_row, 1).fill = S["section_fill"]
    for c in range(1, ncols + 1):
        ws.cell(start_row, c).fill = S["section_fill"]
        ws.cell(start_row, c).font = S["font_section"]
    r = start_row + 1
    ws.cell(r, 1, "Item").font = S["font_h"]
    ws.cell(r, 1).fill = S["navy_fill"]
    ws.cell(r, 2, "Value").font = S["font_h"]
    ws.cell(r, 2).fill = S["navy_fill"]
    r += 1
    for lab, val, fmt, is_input in rows:
        ws.cell(r, 1, lab).font = S["font"]
        ws.cell(r, 1).border = S["thin"]
        c = ws.cell(r, 2, val)
        c.font = S["font"]
        c.border = S["thin"]
        c.alignment = S["right"]
        if isinstance(val, (int, float)) and fmt:
            c.number_format = fmt
        if is_input:
            c.fill = S["input_fill"]
        r += 1
    return r + 1


def _dump_md_table(ws, start_row: int, title: str, tbl: dict, S) -> int:
    headers = tbl.get("headers") or []
    rows = tbl.get("rows") or []
    width = max(len(headers), 2)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=width)
    ws.cell(start_row, 1, title).font = S["font_section"]
    for c in range(1, width + 1):
        ws.cell(start_row, c).fill = S["section_fill"]
        ws.cell(start_row, c).font = S["font_section"]
    r = start_row + 1
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(r, i, h)
        cell.font = S["font_h"]
        cell.fill = S["navy_fill"]
        cell.alignment = S["center"]
        cell.border = S["thin"]
    r += 1
    for row in rows:
        for i, raw in enumerate(row, start=1):
            num = parse_cell_number(raw) if i > 1 else None
            cell = ws.cell(r, i, num if num is not None else raw)
            cell.font = S["font"]
            cell.border = S["thin"]
            if num is not None:
                cell.alignment = S["right"]
                if abs(num) < 2 and ("%" in raw or "margin" in (headers[0] if headers else "").lower()):
                    cell.number_format = _FMT_PCT
                elif raw.strip().startswith("$") or abs(num) >= 5:
                    cell.number_format = _FMT_MM if abs(num) >= 20 else _FMT_SH
            if r % 2 == 0:
                cell.fill = S["alt_fill"]
        r += 1
    return r + 1


def _section_title(ws, row: int, col: int, ncols: int, title: str, S) -> None:
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + ncols - 1)
    cell = ws.cell(row, col, title)
    cell.font = S["font_section"]
    cell.fill = S["section_fill"]
    cell.alignment = S["left"]
    for c in range(col, col + ncols):
        ws.cell(row, c).fill = S["section_fill"]
        ws.cell(row, c).font = S["font_section"]


def _put_input(ws, row, col, value, fmt, S, *, input_cell=True):
    c = ws.cell(row, col, value)
    c.font = S["font"]
    c.alignment = S["right"]
    c.border = S["thin"]
    if fmt:
        c.number_format = fmt
    if input_cell:
        c.fill = S["input_fill"]
    return c


def _put_formula(ws, row, col, formula, fmt, S, *, gold=False, bold=False):
    c = ws.cell(row, col, formula)
    c.font = S["font_bold"] if bold else S["font"]
    c.alignment = S["right"]
    c.border = S["thin"]
    if fmt:
        c.number_format = fmt
    if gold:
        c.fill = S["pale_gold"]
        c.font = S["font_kpi"]
    return c


def _paint_verdict_cell(cell, tone: Optional[str], intensity: Optional[float], S) -> None:
    """Solid fill + readable font from the palette. Never white-on-white."""
    from openpyxl.styles import Font, PatternFill, Alignment
    fill_hex, font_hex = verdict_palette(tone, intensity)
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.font = Font(name="Calibri", size=11, bold=True, color=font_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = S["thin"]


def _apply_verdict_cf(ws, ref: str) -> None:
    """Live UNDER/FAIR/OVER on a formula cell. Dark text on light fills."""
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Font, PatternFill, Alignment
    # Default (pre-calc / unmatched): slate, dark ink — readable on a white sheet.
    cell = ws[ref]
    cell.fill = PatternFill("solid", fgColor="E2E8F0")
    cell.font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rules = [
        ('"UNDERVALUED"', "D1FAE5", "14532D"),
        ('"FAIR VALUED"', "E2E8F0", "0F172A"),
        ('"OVERVALUED"', "FEE2E2", "7F1D1D"),
    ]
    for lit, fill, ink in rules:
        ws.conditional_formatting.add(ref, FormulaRule(
            formula=[f"{ref}={lit}"],
            fill=PatternFill("solid", fgColor=fill),
            font=Font(name="Calibri", size=12, bold=True, color=ink),
        ))


def _write_row3_summary(ws, approaches: list[dict], dcf_ref: str, last_ref: str, S) -> None:
    """Row 3: one pair (name, verdict) per approach. High-contrast fills."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    ws.row_dimensions[3].height = 28
    lab = ws.cell(3, 1, "vs last →")
    lab.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    lab.fill = S["navy_fill"]
    lab.alignment = Alignment(horizontal="center", vertical="center")
    lab.border = S["thin"]
    shown = [a for a in (approaches or []) if a.get("id") != "report_pt"][:6]
    if not shown:
        shown = [{"id": "dcf", "name": "DCF", "tone": None, "intensity": 0, "verdict": None}]
    vf = _dcf_verdict_formulas(dcf_ref, last_ref)
    for i, ap in enumerate(shown):
        name_col = 2 + i * 2
        verd_col = name_col + 1
        nc = ws.cell(3, name_col, ap.get("name") or ap.get("id") or "")
        nc.font = Font(name="Calibri", size=8, bold=True, color=WHITE)
        nc.fill = S["section_fill"]
        nc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        nc.border = S["thin"]
        vc = ws.cell(3, verd_col)
        vc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        vc.border = S["thin"]
        if ap.get("id") == "dcf":
            vc.value = vf["label"]
            _apply_verdict_cf(ws, f"{get_column_letter(verd_col)}3")
        else:
            vc.value = ap.get("verdict") or "—"
            _paint_verdict_cell(vc, ap.get("tone"), ap.get("intensity"), S)
        ws.column_dimensions[get_column_letter(name_col)].width = max(
            12, ws.column_dimensions[get_column_letter(name_col)].width or 12
        )
        ws.column_dimensions[get_column_letter(verd_col)].width = max(
            16, ws.column_dimensions[get_column_letter(verd_col)].width or 16
        )


def _write_approaches_table(
    ws, start_row: int, approaches: list[dict],
    dcf_ref: str, last_ref: str, S,
) -> int:
    """Stacked table: each method separately. Weights are notes only."""
    from openpyxl.formatting.rule import ColorScaleRule
    _section_title(
        ws, start_row, 1, 6,
        "APPROACHES vs LAST  ·  each method on its own (not a weighted blend)",
        S,
    )
    headers = ["Approach", "$ / share", "vs last $", "Gap %", "Verdict", "Notes"]
    hr = start_row + 1
    for i, h in enumerate(headers, start=1):
        c = ws.cell(hr, i, h)
        c.font = S["font_h"]
        c.fill = S["navy_fill"]
        c.alignment = S["center"]
        c.border = S["thin"]
    r = hr + 1
    first = r
    if not approaches:
        ws.cell(r, 1, "No DCF / comps / street / scenario targets parsed.").font = S["font_muted"]
        return r + 2
    for ap in approaches:
        ws.cell(r, 1, ap.get("name") or ap.get("id") or "").font = S["font_bold"]
        ws.cell(r, 1).border = S["thin"]
        is_dcf = ap.get("id") == "dcf"
        if is_dcf:
            _put_formula(ws, r, 2, f"=IF(OR({dcf_ref}=\"nm\",{dcf_ref}=\"\"),\"nm\",{dcf_ref})", _FMT_SH, S, gold=True, bold=True)
        else:
            c = ws.cell(r, 2, ap.get("value"))
            c.number_format = _FMT_SH
            c.font = S["font_bold"]
            c.border = S["thin"]
            c.alignment = S["right"]
        val_ref = f"B{r}"
        _put_formula(
            ws, r, 3,
            f'=IF(OR({val_ref}="nm",{val_ref}="",{last_ref}="",{last_ref}=0),"—",{val_ref}-{last_ref})',
            _FMT_SH, S,
        )
        _put_formula(
            ws, r, 4,
            f'=IF(OR({val_ref}="nm",{val_ref}="",{last_ref}="",{last_ref}=0),"—",{val_ref}/{last_ref}-1)',
            _FMT_PCT, S,
        )
        if is_dcf:
            live = _dcf_verdict_formulas(val_ref, last_ref)
            _put_formula(ws, r, 5, live["label"], None, S, bold=True)
            from openpyxl.utils import get_column_letter
            _apply_verdict_cf(ws, f"E{r}")
        else:
            vc = ws.cell(r, 5, ap.get("verdict") or "—")
            _paint_verdict_cell(vc, ap.get("tone"), ap.get("intensity"), S)
        note = ws.cell(r, 6, ap.get("note") or "")
        note.font = S["font_muted"]
        note.border = S["thin"]
        r += 1
    last = r - 1
    if last >= first:
        ws.conditional_formatting.add(
            f"D{first}:D{last}",
            ColorScaleRule(
                start_type="num", start_value=-0.40, start_color="B91C1C",
                mid_type="num", mid_value=0.0, mid_color="E2E8F0",
                end_type="num", end_value=0.40, end_color="047857",
            ),
        )
    foot = ws.cell(
        r, 1,
        "Not blended. Green = cheap vs last (undervalued). Red = rich vs last (overvalued). "
        "Fair = within ±5%. Report weights are shown as notes only.",
    )
    foot.font = S["font_muted"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    return r + 2


def _build_dcf_horizon(
    financials: dict,
    forecasts: dict[int, dict[str, float]],
    ladder: list[dict],
    n: int = 5,
) -> list[dict[str, Any]]:
    """Year 0 actual + n explicit forecast years for the DCF ladder / pro forma."""
    annuals = _annuals_oldest_first(financials)
    last = annuals[-1] if annuals else {}
    try:
        last_fy = int(last["fy"]) if last.get("fy") is not None else None
    except (TypeError, ValueError):
        last_fy = None
    by_t = {int(r["t"]): r for r in ladder if r.get("t") is not None}
    by_fy = {int(r["fy"]): r for r in ladder if r.get("fy") is not None}

    def _rev(src, period=None, already=True):
        if src and src.get("revenue") is not None:
            return src["revenue"]
        if src and src.get("Revenue") is not None:
            return to_model_units("Revenue", src["Revenue"], already_millions=already)
        if period:
            return _period_money(period, "Revenue", already_millions=already)
        return None

    def _fcf(src, period=None, already=True):
        if src and src.get("fcf") is not None:
            return src["fcf"]
        if src and src.get("FreeCashFlow") is not None:
            return to_model_units("FreeCashFlow", src["FreeCashFlow"], already_millions=already)
        if period:
            return _period_money(period, "FreeCashFlow", already_millions=already)
        return None

    years: list[dict[str, Any]] = []
    y0 = {
        "t": 0,
        "fy": last_fy,
        "kind": "A",
        "revenue": _rev(by_t.get(0) or by_fy.get(last_fy or -1), last, already=False),
        "fcf": _fcf(by_t.get(0) or by_fy.get(last_fy or -1), last, already=False),
        "ebit": _period_money(last, "OperatingIncome"),
        "ebitda": _period_money(last, "EBITDA"),
        "ni": _period_money(last, "NetIncome"),
        "eps": _f(last.get("DilutedEPS")),
        "gp": _period_money(last, "GrossProfit"),
        "hard_fcf": True,
        "hard_rev": True,
    }
    years.append(y0)
    y0_rev = y0.get("revenue")
    y0_fcf = y0.get("fcf")
    for t in range(1, n + 1):
        fy = (last_fy + t) if last_fy else None
        src_l = by_t.get(t) or (by_fy.get(fy) if fy else None) or {}
        src_f = forecasts.get(fy) or {}
        rev = reconcile_money_level(
            _rev(src_l, None, already=True) or _rev(src_f, None, already=True),
            y0_rev,
        )
        fcf = reconcile_money_level(
            _fcf(src_l, None, already=True) or _fcf(src_f, None, already=True),
            y0_fcf,
        )
        item = {
            "t": t,
            "fy": fy,
            "kind": "E",
            "revenue": rev,
            "fcf": fcf,
            "ebit": to_model_units("OperatingIncome", src_f.get("OperatingIncome"), already_millions=True)
                    if src_f else None,
            "ebitda": to_model_units("EBITDA", src_f.get("EBITDA"), already_millions=True) if src_f else None,
            "ni": to_model_units("NetIncome", src_f.get("NetIncome"), already_millions=True) if src_f else None,
            "eps": _f(src_f.get("DilutedEPS")) if src_f else None,
            "gp": to_model_units("GrossProfit", src_f.get("GrossProfit"), already_millions=True) if src_f else None,
            "rev_g_in": _as_rate(src_l.get("rev_g")) or _as_rate(src_f.get("RevenueGrowth")),
            "fcf_g_in": _as_rate(src_l.get("fcf_g")) or _as_rate(src_f.get("FCFGrowth")),
        }
        item["hard_fcf"] = item["fcf"] is not None
        item["hard_rev"] = item["revenue"] is not None
        years.append(item)
    return years


def _dcf_verdict_formulas(dcf_ref: str, last_ref: str, band: float = _FAIR_BAND) -> dict[str, str]:
    """Live Excel formulas: UNDERVALUED / FAIR VALUED / OVERVALUED vs last price."""
    label = (
        f'IF(OR({dcf_ref}="nm",{dcf_ref}="",{last_ref}="",{last_ref}=0),"—",'
        f'IF(ABS({dcf_ref}/{last_ref}-1)<{band},"FAIR VALUED",'
        f'IF({dcf_ref}>{last_ref},"UNDERVALUED","OVERVALUED")))'
    )
    byline = (
        f'IF(OR({dcf_ref}="nm",{last_ref}="",{last_ref}=0),"—",'
        f'IF(ABS({dcf_ref}/{last_ref}-1)<{band},'
        f'"within ±5%  ·  "&TEXT({dcf_ref}-{last_ref},"$#,##0.00")&"/sh",'
        f'TEXT(ABS({dcf_ref}/{last_ref}-1),"0.0%")'
        f'&IF({dcf_ref}>{last_ref}," cheap  ·  "," rich  ·  ")'
        f'&TEXT({dcf_ref}-{last_ref},"$#,##0.00")&"/sh"))'
    )
    gap_pct = (
        f'IF(OR({dcf_ref}="nm",{last_ref}="",{last_ref}=0),"—",{dcf_ref}/{last_ref}-1)'
    )
    gap_px = (
        f'IF(OR({dcf_ref}="nm",{last_ref}="",{last_ref}=0),"—",{dcf_ref}-{last_ref})'
    )
    return {
        "label": "=" + label,
        "byline": "=" + byline,
        "gap_pct": "=" + gap_pct,
        "gap_px": "=" + gap_px,
    }


def _dcf_base_from_approaches(approaches: Optional[list]) -> Optional[float]:
    """PT-derivation DCF (base) implied $/share — not the live model DCF."""
    for a in approaches or []:
        if a.get("id") == "dcf_base":
            v = _f(a.get("value"))
            if v is not None and v > 1:
                return v
    return None


def _write_dcf_base_bridge(
    ws,
    start_row: int,
    *,
    dcf_base: Optional[float],
    sum_row: int,
    last_fcf_row: int,
    model_dcf_ref: str,
    S,
) -> None:
    """Reverse-engineer report DCF (base) in columns E–G, beside the Gordon bridge.

    Holds the model’s FCF ladder, shares, and net debt fixed so the gap versus
    live DCF value/share shows up in implied TV / implied g.
    """
    c0, c1, c2 = 5, 6, 7
    _section_title(
        ws, start_row, c0, 3,
        "DCF (BASE) BRIDGE  ·  reverse from the report implied — not the live Gordon walk",
        S,
    )
    hdr = start_row + 1
    for col, h in ((c0, "Step"), (c1, "Value"), (c2, "Note")):
        cell = ws.cell(hdr, col, h)
        cell.font = S["font_h"]
        cell.fill = S["navy_fill"]
        cell.alignment = S["center"]
        cell.border = S["thin"]

    r = hdr + 1

    def put(lab, val, fmt, note, *, input_cell=False, gold=False) -> int:
        nonlocal r
        ws.cell(r, c0, lab).font = S["font_bold"] if gold else S["font"]
        ws.cell(r, c0).border = S["thin"]
        if isinstance(val, str) and str(val).startswith("="):
            _put_formula(ws, r, c1, val, fmt, S, gold=gold, bold=gold)
        else:
            _put_input(ws, r, c1, val, fmt, S, input_cell=input_cell)
            if gold and not input_cell:
                ws.cell(r, c1).fill = S["pale_gold"]
                ws.cell(r, c1).font = S["font_kpi"]
        n = ws.cell(r, c2, note)
        n.font = S["font_muted"]
        n.border = S["thin"]
        n.alignment = Alignment_wrap(S)
        this = r
        r += 1
        return this

    r_px = put(
        "DCF (base) $/share", dcf_base, _FMT_SH,
        "PT-derivation implied — yellow input", input_cell=True,
    )
    r_sh = put("× Diluted shares (m)", "=$B$20", _FMT_SHARES,
               "same share count as the model bridge")
    r_eq = put(
        "= Implied equity ($m)",
        f'=IF(OR(F{r_px}="",F{r_sh}=""),"nm",F{r_px}*F{r_sh})',
        _FMT_MM, "DCF (base) × shares",
    )
    r_nd = put("(+) Net debt / (−) cash", "=$B$19", _FMT_MM,
               "same BS plug as the model bridge")
    r_ev = put(
        "= Implied EV ($m)",
        f'=IF(F{r_eq}="nm","nm",F{r_eq}+F{r_nd})',
        _FMT_MM, "equity + net debt", gold=True,
    )
    r_pv = put("(−) PV of explicit FCF", f"=H{sum_row}", _FMT_MM,
               "model ladder held constant — the gap sits in TV")
    r_pvtv = put(
        "= Implied PV of TV ($m)",
        f'=IF(F{r_ev}="nm","nm",F{r_ev}-F{r_pv})',
        _FMT_MM, "implied EV − Σ PV FCF",
    )
    r_df = put("÷ Year-n discount factor", f"=G{last_fcf_row}", "0.000",
               "1/(1+WACC)^n from the ladder")
    r_tv = put(
        "= Implied TV ($m)",
        f'=IF(OR(F{r_pvtv}="nm",F{r_df}="",F{r_df}=0),"nm",F{r_pvtv}/F{r_df})',
        _FMT_MM, "undiscounted TV that hits DCF (base)", gold=True,
    )
    r_fcf = put("Year-n FCF ($m)", f"=E{last_fcf_row}", _FMT_MM,
                "same FCFn as the model bridge")
    put(
        "Implied TV / FCFn",
        f'=IF(OR(F{r_tv}="nm",F{r_fcf}="",F{r_fcf}=0),"nm",F{r_tv}/F{r_fcf})',
        _FMT_X, "exit multiple implied by DCF (base)",
    )
    r_g = put(
        "Implied g (Gordon)",
        f'=IF(OR(F{r_tv}="nm",F{r_tv}="",F{r_fcf}="",F{r_fcf}=0,$B$15=""),"nm",'
        f'(F{r_tv}*$B$15-F{r_fcf})/(F{r_tv}+F{r_fcf}))',
        _FMT_PCT,
        "g that would make this sheet's ladder equal DCF (base)",
    )
    r_gm = put("Model g", "=$B$16", _FMT_PCT, "from the WACC build")
    put(
        "Δ g (implied − model)",
        f'=IF(OR(F{r_g}="nm",F{r_gm}=""),"nm",F{r_g}-F{r_gm})',
        _FMT_PCT,
        "negative = report haircut the terminal (conservatism vs live model)",
    )

    _section_title(ws, r, c0, 3, "GAP  ·  live model DCF vs report DCF (base)", S)
    r += 1
    r_mod = put(
        "Model DCF value / share", f"={model_dcf_ref}", _FMT_SH,
        "live Gordon / ladder / this sheet", gold=True,
    )
    r_base = put("DCF (base) $/share", f"=F{r_px}", _FMT_SH,
                 "report PT-derivation implied")
    put(
        "$ gap (model − base)",
        f'=IF(OR(F{r_mod}="nm",F{r_base}="",F{r_base}=0),"nm",F{r_mod}-F{r_base})',
        _FMT_SH,
        "positive = model above the report DCF (base)", gold=True,
    )
    put(
        "% gap",
        f'=IF(OR(F{r_mod}="nm",F{r_base}="",F{r_base}=0),"nm",F{r_mod}/F{r_base}-1)',
        _FMT_PCT,
        "model is unadjusted Gordon; DCF (base) is the report implied, often a haircut",
        gold=True,
    )
    ws.row_dimensions[r - 1].height = 28


def _dcf_price_formula(wacc_ref: str, g_ref: str, fcf_refs: list[str],
                       nd_ref: str, sh_ref: str) -> str:
    """Equity value / share at a given WACC and g, using explicit FCF years 1..n."""
    n = len(fcf_refs)
    pvs = "+".join(f"{fcf_refs[i]}/(1+{wacc_ref})^{i+1}" for i in range(n))
    last = fcf_refs[-1]
    tv = f"{last}*(1+{g_ref})/({wacc_ref}-{g_ref})/(1+{wacc_ref})^{n}"
    return (
        f'IF(OR({sh_ref}=0,{sh_ref}="",{wacc_ref}<={g_ref}),"nm",'
        f"({pvs}+{tv}-{nd_ref})/{sh_ref})"
    )


def _valuation_sheet(
    wb,
    *,
    capital: dict,
    dcf: dict,
    wacc_build: dict,
    ladder: list,
    sensitivity: Optional[dict],
    forecasts: dict,
    financials: dict,
    derivation,
    comps,
    bridge_tbl,
    pt,
    price,
    S,
    approaches: Optional[list] = None,
):
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Valuation")
    _write_banner(
        ws, 12,
        "VALUATION  ·  DCF BUILD",
        "Yellow = inputs  ·  Formulas audit WACC, ladder, terminal value, and the WACC × g grid  ·  $ millions except per-share",
        S,
    )

    merged = {**dcf, **wacc_build}
    horizon = _build_dcf_horizon(financials, forecasts, ladder, n=5)
    n = max((y["t"] for y in horizon), default=0)

    # Seed capital-structure weights from the BS if the report omitted them.
    nd = merged.get("net_debt")
    if nd is None:
        nd = capital.get("net_debt")
    shares = normalize_shares_millions(merged.get("shares"), capital.get("shares"))
    mkt = capital.get("market_cap")
    we = merged.get("we")
    wd = merged.get("wd")
    if we is None and wd is not None:
        we = 1.0 - wd
    if we is None and mkt and nd is not None:
        ev = mkt + nd
        if ev:
            we = max(0.0, min(1.0, mkt / ev))
    if we is None:
        we = 0.85
    if wd is None:
        wd = 1.0 - we

    rf = merged.get("rf") if merged.get("rf") is not None else 0.04
    erp = merged.get("erp") if merged.get("erp") is not None else 0.05
    beta = merged.get("beta") if merged.get("beta") is not None else 1.0
    kd = merged.get("kd_pretax") if merged.get("kd_pretax") is not None else 0.05
    tax = merged.get("tax_rate") if merged.get("tax_rate") is not None else 0.21
    g = merged.get("terminal_growth") if merged.get("terminal_growth") is not None else 0.025
    # FCF growth used only to fill missing explicit years.
    known_fcf = [y for y in horizon if y.get("fcf") is not None]
    fcf_g = 0.08
    if len(known_fcf) >= 2:
        a, b = known_fcf[0]["fcf"], known_fcf[1]["fcf"]
        if a:
            fcf_g = b / a - 1.0
    elif horizon[0].get("fcf") and len(horizon) > 1 and horizon[1].get("fcf"):
        if horizon[0]["fcf"]:
            fcf_g = horizon[1]["fcf"] / horizon[0]["fcf"] - 1.0
    else:
        g_in = next((y.get("fcf_g_in") for y in horizon if y.get("fcf_g_in") is not None), None)
        if g_in is not None and -0.5 <= g_in <= 1.5:
            fcf_g = g_in
    rev_g = 0.08
    known_rev = [y for y in horizon if y.get("revenue") is not None]
    if len(known_rev) >= 2 and known_rev[0]["revenue"]:
        rev_g = known_rev[1]["revenue"] / known_rev[0]["revenue"] - 1.0
    else:
        g_in = next((y.get("rev_g_in") for y in horizon if y.get("rev_g_in") is not None), None)
        if g_in is not None and -0.5 <= g_in <= 0.8:
            rev_g = g_in

    # ── WACC BUILD (A–C) ────────────────────────────────────────────────
    _section_title(ws, 4, 1, 3, "WACC BUILD", S)
    ws.cell(5, 1, "Component").font = S["font_h"]
    ws.cell(5, 2, "Value").font = S["font_h"]
    ws.cell(5, 3, "Notes").font = S["font_h"]
    for c in range(1, 4):
        ws.cell(5, c).fill = S["navy_fill"]
        ws.cell(5, c).font = S["font_h"]
        ws.cell(5, c).border = S["thin"]

    wacc_rows = [
        (6, "Risk-free rate", rf, _FMT_PCT, True, merged.get("rf_notes") or "10y / source in report"),
        (7, "Equity risk premium", erp, _FMT_PCT, True, merged.get("erp_notes") or ""),
        (8, "Beta (levered)", beta, "0.00", True, merged.get("beta_notes") or ""),
        (9, "Cost of equity", None, _FMT_PCT, False, "rf + β × ERP"),
        (10, "Pre-tax cost of debt", kd, _FMT_PCT, True, merged.get("kd_pretax_notes") or ""),
        (11, "Tax rate", tax, _FMT_PCT, True, merged.get("tax_rate_notes") or ""),
        (12, "After-tax cost of debt", None, _FMT_PCT, False, "kd × (1 − tax)"),
        (13, "E / (D+E)", we, _FMT_PCT, True, merged.get("we_notes") or "equity weight"),
        (14, "D / (D+E)", None, _FMT_PCT, False, "1 − E/(D+E)"),
        (15, "WACC (base)", None, _FMT_PCT, False, "ke×we + kd_at×wd"),
        (16, "Terminal growth (g)", g, _FMT_PCT, True, merged.get("terminal_growth_notes") or "long-run GDP / inflation"),
        (17, "FCF growth (missing yrs)", fcf_g, _FMT_PCT, True, "used only when a forecast year has no FCF"),
        (18, "Revenue growth (missing yrs)", rev_g, _FMT_PCT, True, "used only when a forecast year has no revenue"),
        (19, "Net debt / (cash) ($m)", nd if nd is not None else 0.0, _FMT_MM, True, "latest BS; cash-rich → negative"),
        (20, "Diluted shares (m)", shares, _FMT_SHARES, True, merged.get("shares_notes") or ""),
        (21, "Last price", price, _FMT_SH, True, ""),
        (22, "Report 12m PT", pt, _FMT_SH, False, ""),
    ]
    for row, lab, val, fmt, is_in, notes in wacc_rows:
        ws.cell(row, 1, lab).font = S["font_bold"] if row in (15, 16) else S["font"]
        ws.cell(row, 1).border = S["thin"]
        ws.cell(row, 3, notes).font = S["font_muted"]
        ws.cell(row, 3).border = S["thin"]
        if row == 9:
            _put_formula(ws, row, 2, "=B6+B8*B7", fmt, S)
        elif row == 12:
            _put_formula(ws, row, 2, "=B10*(1-B11)", fmt, S)
        elif row == 14:
            _put_formula(ws, row, 2, "=1-B13", fmt, S)
        elif row == 15:
            _put_formula(ws, row, 2, "=B9*B13+B12*B14", fmt, S, gold=True, bold=True)
        else:
            _put_input(ws, row, 2, val, fmt, S, input_cell=is_in)

    # ── PRO FORMA YEARS (E–L) ───────────────────────────────────────────
    pf_start_col = 5  # E
    _section_title(ws, 4, pf_start_col, max(len(horizon) + 1, 4), "PRO FORMA YEARS  ($ millions)", S)
    ws.cell(5, pf_start_col, "Metric").font = S["font_h"]
    ws.cell(5, pf_start_col).fill = S["navy_fill"]
    ws.cell(5, pf_start_col).border = S["thin"]
    pf_cols = []  # (excel_col, year_dict)
    for i, y in enumerate(horizon):
        col = pf_start_col + 1 + i
        fy = y.get("fy")
        lab = f"FY{fy}{'A' if y['kind']=='A' else 'E'}" if fy else f"Y{y['t']}"
        cell = ws.cell(5, col, lab)
        cell.font = S["font_h"]
        cell.fill = S["gold_fill"] if y["kind"] == "E" else S["navy_fill"]
        cell.alignment = S["center"]
        cell.border = S["thin"]
        pf_cols.append((col, y))

    pf_metrics = [
        (6, "Revenue", "revenue", _FMT_MM),
        (7, "  YoY growth", "rev_g", _FMT_PCT),
        (8, "Gross profit", "gp", _FMT_MM),
        (9, "Operating income", "ebit", _FMT_MM),
        (10, "  Operating margin", "om", _FMT_PCT),
        (11, "EBITDA", "ebitda", _FMT_MM),
        (12, "Net income", "ni", _FMT_MM),
        (13, "Diluted EPS", "eps", _FMT_SH),
        (14, "Free cash flow", "fcf", _FMT_MM),
        (15, "  FCF growth", "fcf_g", _FMT_PCT),
        (16, "  FCF margin", "fcf_m", _FMT_PCT),
    ]
    for row, lab, key, fmt in pf_metrics:
        ws.cell(row, pf_start_col, lab).font = S["font"]
        ws.cell(row, pf_start_col).border = S["thin"]
        for i, (col, y) in enumerate(pf_cols):
            letter = get_column_letter(col)
            prev = get_column_letter(pf_cols[i - 1][0]) if i else None
            cell = ws.cell(row, col)
            cell.border = S["thin"]
            cell.alignment = S["right"]
            cell.number_format = fmt
            if key == "rev_g":
                if prev:
                    cell.value = f'=IF(OR({prev}6="",{prev}6=0),"—",({letter}6-{prev}6)/{prev}6)'
                    cell.font = S["font"]
                else:
                    cell.value = "—"
                    cell.font = S["font_muted"]
            elif key == "om":
                cell.value = f'=IF(OR({letter}6="",{letter}6=0),"—",{letter}9/{letter}6)'
                cell.font = S["font"]
            elif key == "fcf_g":
                if prev:
                    cell.value = f'=IF(OR({prev}14="",{prev}14=0),"—",({letter}14-{prev}14)/{prev}14)'
                    cell.font = S["font"]
                else:
                    cell.value = "—"
                    cell.font = S["font_muted"]
            elif key == "fcf_m":
                cell.value = f'=IF(OR({letter}6="",{letter}6=0),"—",{letter}14/{letter}6)'
                cell.font = S["font"]
            elif key == "revenue":
                if y.get("hard_rev") and y.get("revenue") is not None:
                    cell.value = y["revenue"]
                    cell.fill = S["input_fill"] if y["kind"] == "E" else S["white_fill"]
                    cell.font = S["font_est"] if y["kind"] == "E" else S["font"]
                elif prev:
                    cell.value = f"={prev}6*(1+$B$18)"
                    cell.font = S["font_est"]
                else:
                    cell.value = y.get("revenue")
                    cell.font = S["font"]
            elif key == "fcf":
                if y.get("hard_fcf") and y.get("fcf") is not None:
                    cell.value = y["fcf"]
                    cell.fill = S["input_fill"] if y["kind"] == "E" else S["white_fill"]
                    cell.font = S["font_est"] if y["kind"] == "E" else S["font"]
                elif prev:
                    cell.value = f"={prev}14*(1+$B$17)"
                    cell.font = S["font_est"]
                else:
                    cell.value = y.get("fcf")
                    cell.font = S["font"]
            else:
                cell.value = y.get(key)
                cell.font = S["font_est"] if y["kind"] == "E" else S["font"]
                if y["kind"] == "E" and y.get(key) is not None:
                    cell.fill = S["input_fill"]

    # ── DCF PROJECTION LADDER ───────────────────────────────────────────
    lad_row = 24
    _section_title(ws, lad_row, 1, 8, "DCF PROJECTION LADDER (BASE CASE)  ·  FCFF, year-end discounting", S)
    lad_headers = [
        "Year", "Fiscal", "Revenue ($m)", "Rev growth",
        "FCF ($m)", "FCF growth", "Discount factor", "PV of FCF ($m)",
    ]
    hdr = lad_row + 1
    for i, h in enumerate(lad_headers, start=1):
        c = ws.cell(hdr, i, h)
        c.font = S["font_h"]
        c.fill = S["navy_fill"]
        c.alignment = S["center"]
        c.border = S["thin"]

    # Map t → pro forma column letter (revenue row 6, fcf row 14)
    pf_by_t = {y["t"]: get_column_letter(col) for col, y in pf_cols}
    fcf_cells: list[str] = []  # years 1..n absolute refs
    first_exp = hdr + 2  # year 1 row
    last_exp = hdr + 1 + n
    for y in horizon:
        r = hdr + 1 + y["t"]
        t = y["t"]
        ws.cell(r, 1, f"{t} ({y['kind']})").font = S["font_bold"]
        ws.cell(r, 1).border = S["thin"]
        fy = y.get("fy")
        ws.cell(r, 2, f"FY{fy}" if fy else "").border = S["thin"]
        ws.cell(r, 2).alignment = S["center"]
        pf_let = pf_by_t.get(t)
        # Revenue / FCF linked to pro forma so one set of inputs drives both
        if pf_let:
            _put_formula(ws, r, 3, f"={pf_let}6", _FMT_MM, S)
            _put_formula(ws, r, 5, f"={pf_let}14", _FMT_MM, S)
        else:
            _put_input(ws, r, 3, y.get("revenue"), _FMT_MM, S, input_cell=y["kind"] == "E")
            _put_input(ws, r, 5, y.get("fcf"), _FMT_MM, S, input_cell=y["kind"] == "E")
        if t == 0:
            ws.cell(r, 4, "—").border = S["thin"]
            ws.cell(r, 6, "—").border = S["thin"]
            _put_formula(ws, r, 7, 1, "0.000", S)
            ws.cell(r, 8, "—").border = S["thin"]
        else:
            prev_r = r - 1
            _put_formula(ws, r, 4, f'=IF(OR(C{prev_r}="",C{prev_r}=0),"—",(C{r}-C{prev_r})/C{prev_r})', _FMT_PCT, S)
            _put_formula(ws, r, 6, f'=IF(OR(E{prev_r}="",E{prev_r}=0),"—",(E{r}-E{prev_r})/E{prev_r})', _FMT_PCT, S)
            _put_formula(ws, r, 7, f"=1/(1+$B$15)^{t}", "0.000", S)
            _put_formula(ws, r, 8, f"=E{r}*G{r}", _FMT_MM, S)
            fcf_cells.append(f"$E${r}")
        for c in range(1, 9):
            ws.cell(r, c).border = S["thin"]

    sum_row = hdr + 2 + n
    ws.cell(sum_row, 1, "Sum PV explicit FCF").font = S["font_bold"]
    ws.cell(sum_row, 1).fill = S["pale_navy"]
    for c in range(1, 8):
        ws.cell(sum_row, c).fill = S["pale_navy"]
        ws.cell(sum_row, c).border = S["thin"]
    _put_formula(ws, sum_row, 8, f"=SUM(H{first_exp}:H{last_exp})", _FMT_MM, S, gold=True, bold=True)

    # ── TERMINAL VALUE & EQUITY BRIDGE ──────────────────────────────────
    br = sum_row + 2
    _section_title(ws, br, 1, 3, "TERMINAL VALUE & EQUITY BRIDGE", S)
    last_fcf_row = last_exp
    bridge = [
        (br + 1, "Year-n FCF ($m)", f"=E{last_fcf_row}", _FMT_MM, False),
        (br + 2, "Terminal growth g", "=$B$16", _FMT_PCT, False),
        (br + 3, "WACC", "=$B$15", _FMT_PCT, False),
        (br + 4, "Terminal value (Gordon)", f"=IF(B{br+3}<=B{br+2},\"nm\",B{br+1}*(1+B{br+2})/(B{br+3}-B{br+2}))", _FMT_MM, False),
        (br + 5, "PV of terminal value", f"=IF(OR(B{br+4}=\"nm\",B{br+4}=\"\"),\"nm\",B{br+4}*G{last_fcf_row})", _FMT_MM, False),
        (br + 6, "Enterprise value", f"=H{sum_row}+IF(B{br+5}=\"nm\",0,B{br+5})", _FMT_MM, True),
        (br + 7, "(−) Net debt / (+) net cash", "=$B$19", _FMT_MM, False),
        (br + 8, "Equity value", f"=B{br+6}-B{br+7}", _FMT_MM, True),
        (br + 9, "Diluted shares (m)", "=$B$20", _FMT_SHARES, False),
        (br + 10, "DCF value / share", f"=IF(B{br+9}=0,\"nm\",B{br+8}/B{br+9})", _FMT_SH, True),
        (br + 11, "Last price", "=$B$21", _FMT_SH, False),
        (br + 12, "Upside / (downside)", f"=IF(OR(B{br+11}=0,B{br+10}=\"nm\"),\"nm\",B{br+10}/B{br+11}-1)", _FMT_PCT, True),
        (br + 13, "Report 12m PT", "=$B$22", _FMT_SH, False),
        (br + 14, "DCF vs report PT", f"=IF(OR(B{br+13}=0,B{br+10}=\"nm\"),\"nm\",B{br+10}/B{br+13}-1)", _FMT_PCT, False),
    ]
    dcf_ref = f"B{br+10}"
    last_ref = "$B$21"
    vf = _dcf_verdict_formulas(dcf_ref, last_ref)
    bridge.extend([
        (br + 15, "DCF verdict (vs last)", vf["label"], None, True),
        (br + 16, "Mispricing vs last", vf["gap_pct"], _FMT_PCT, True),
        (br + 17, "$ / sh vs last", vf["gap_px"], _FMT_SH, False),
    ])
    for row, lab, formula, fmt, gold in bridge:
        ws.cell(row, 1, lab).font = S["font_bold"] if gold else S["font"]
        ws.cell(row, 1).border = S["thin"]
        _put_formula(ws, row, 2, formula, fmt, S, gold=gold, bold=gold)
        ws.cell(row, 3, "").border = S["thin"]

    _write_dcf_base_bridge(
        ws, br,
        dcf_base=_dcf_base_from_approaches(approaches),
        sum_row=sum_row,
        last_fcf_row=last_fcf_row,
        model_dcf_ref=dcf_ref,
        S=S,
    )

    # Row-3 summary of every approach (DCF live; others from the report).
    _write_row3_summary(ws, approaches or [], dcf_ref, last_ref, S)
    # Live DCF verdict in the bridge — same high-contrast CF as row 3.
    _apply_verdict_cf(ws, f"B{br + 15}")

    # Gordon note
    note_r = br + 18
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r + 1, end_column=3)
    ws.cell(
        note_r, 1,
        "TV = FCFn × (1+g) / (WACC − g).  Discount factor = 1/(1+WACC)^t (year-end).  "
        "Change yellow cells — ladder, EV, $/share, and the sensitivity grid all recast.",
    ).font = S["font_muted"]
    ws.cell(note_r, 1).alignment = Alignment_wrap(S)

    # ── SENSITIVITY GRID ────────────────────────────────────────────────
    sens_row = 24
    sens_col = 10  # J
    _section_title(ws, sens_row, sens_col, 6, "DCF SENSITIVITY  ($ / share)  ·  WACC × terminal growth", S)
    base_wacc = merged.get("wacc")
    if base_wacc is None:
        # approximate from seeded build-up so the gold cell lands near the formula WACC
        ke = rf + beta * erp
        kd_at = kd * (1 - tax)
        base_wacc = ke * we + kd_at * wd
    base_g = g
    if sensitivity and sensitivity.get("waccs") and sensitivity.get("tgrs"):
        wacc_axis = list(sensitivity["waccs"])
        g_axis = list(sensitivity["tgrs"])
    else:
        wacc_axis = sorted({round(base_wacc + d, 4) for d in (-0.02, -0.01, 0.0, 0.01, 0.02)})
        g_axis = [0.01, 0.02, 0.025, 0.03, 0.035]
        if round(base_g, 4) not in [round(x, 4) for x in g_axis]:
            g_axis = sorted(g_axis + [round(base_g, 4)])

    # Header row: TGR values
    ws.cell(sens_row + 1, sens_col, "WACC \\ g").font = S["font_h"]
    ws.cell(sens_row + 1, sens_col).fill = S["navy_fill"]
    ws.cell(sens_row + 1, sens_col).border = S["thin"]
    for j, gv in enumerate(g_axis):
        c = ws.cell(sens_row + 1, sens_col + 1 + j, gv)
        c.font = S["font_h"]
        c.fill = S["navy_fill"]
        c.number_format = _FMT_PCT
        c.alignment = S["center"]
        c.border = S["thin"]

    nd_ref, sh_ref = "$B$19", "$B$20"
    for i, wv in enumerate(wacc_axis):
        r = sens_row + 2 + i
        wc = ws.cell(r, sens_col, wv)
        wc.number_format = _FMT_PCT
        wc.font = S["font_bold"]
        wc.fill = S["pale_navy"]
        wc.border = S["thin"]
        wacc_cell = f"{get_column_letter(sens_col)}{r}"
        for j, gv in enumerate(g_axis):
            g_cell = f"{get_column_letter(sens_col + 1 + j)}${sens_row + 1}"
            formula = "=" + _dcf_price_formula(wacc_cell, g_cell, fcf_cells, nd_ref, sh_ref)
            cell = ws.cell(r, sens_col + 1 + j, formula)
            cell.number_format = _FMT_SH
            cell.border = S["thin"]
            cell.alignment = S["center"]
            if abs(wv - base_wacc) < 0.0006 and abs(gv - base_g) < 0.0006:
                cell.fill = S["pale_gold"]
                cell.font = S["font_kpi"]

    ws.cell(sens_row + 3 + len(wacc_axis), sens_col,
            "Gold cell = base WACC & g. Grid is live — it uses the same FCF ladder as the bridge.").font = S["font_muted"]
    ws.merge_cells(
        start_row=sens_row + 3 + len(wacc_axis), start_column=sens_col,
        end_row=sens_row + 3 + len(wacc_axis), end_column=sens_col + min(5, len(g_axis)),
    )

    # ── Each valuation approach vs last (not blended) ────────────────────
    ap_row = note_r + 3
    ap_end = _write_approaches_table(
        ws, ap_row, approaches or [], dcf_ref, last_ref, S,
    )

    # ── Trading multiples + comps / derivation below ────────────────────
    tm_row = max(ap_end + 1, br + 17, sens_row + 6 + len(wacc_axis))
    _section_title(ws, tm_row, 1, 3, "TRADING MULTIPLES (LAST REPORTED FY / TTM)", S)
    tm_rows = [
        ("Market cap ($m)", capital.get("market_cap"), _FMT_MM),
        ("Enterprise value ($m)", capital.get("enterprise_value"), _FMT_MM),
        ("P / E", capital.get("pe"), _FMT_X),
        ("EV / EBITDA", capital.get("ev_ebitda"), _FMT_X),
        ("EV / Sales", capital.get("ev_sales"), _FMT_X),
        ("P / B", capital.get("pb"), _FMT_X),
        ("FCF yield", capital.get("fcf_yield"), _FMT_PCT),
    ]
    rr = tm_row + 1
    for lab, val, fmt in tm_rows:
        ws.cell(rr, 1, lab).font = S["font"]
        ws.cell(rr, 1).border = S["thin"]
        c = ws.cell(rr, 2, val)
        c.font = S["font"]
        c.border = S["thin"]
        c.alignment = S["right"]
        if isinstance(val, (int, float)):
            c.number_format = fmt
        rr += 1

    extra = rr + 1
    if derivation:
        extra = _dump_md_table(ws, extra, "PRICE TARGET DERIVATION (FROM REPORT)", derivation, S)
    if comps:
        extra = _dump_md_table(ws, extra, "COMPARABLE COMPANIES (FROM REPORT)", comps, S)
    if bridge_tbl:
        extra = _dump_md_table(ws, extra, "REPORT EQUITY BRIDGE (AS WRITTEN)", bridge_tbl, S)
    if sensitivity and sensitivity.get("table"):
        extra = _dump_md_table(ws, extra, "REPORT SENSITIVITY GRID (AS WRITTEN)", sensitivity["table"], S)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    for i in range(5, 20):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.column_dimensions["E"].width = 34
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 28
    ws.freeze_panes = "A4"
    _print_setup(ws, landscape=True)
    ws.sheet_properties.tabColor = GOLD
    ws.print_title_rows = "1:2"


def _scenarios_sheet(wb, scenarios, price, pt, S):
    ws = wb.create_sheet("Scenarios")
    _write_banner(ws, 6, "SCENARIOS", "Bull / Base / Bear  ·  probability-weighted expected value", S)
    headers = ["Case", "Price target", "Probability", "Implied return", "Contribution to EV"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(4, i, h)
        c.font = S["font_h"]
        c.fill = S["navy_fill"]
        c.alignment = S["center"]
        c.border = S["thin"]
    r = 5
    first_data = r
    if not scenarios:
        ws.cell(r, 1, "No bull / base / bear targets parsed from the report.").font = S["font_muted"]
        r += 1
    else:
        for sc in scenarios:
            ws.cell(r, 1, sc["case"]).font = S["font_bold"]
            ws.cell(r, 1).border = S["thin"]
            c = ws.cell(r, 2, sc.get("price_target"))
            c.number_format = _FMT_SH
            c.font = S["font"]
            c.border = S["thin"]
            p = ws.cell(r, 3, sc.get("probability"))
            p.number_format = _FMT_PCT
            p.fill = S["input_fill"]
            p.border = S["thin"]
            # implied return vs last
            if price and sc.get("price_target") is not None:
                ir = ws.cell(r, 4, f"=IF(OR($B$20=0,$B$20=\"\"),\"—\",B{r}/$B$20-1)")
            else:
                ir = ws.cell(r, 4, None)
            ir.number_format = _FMT_PCT
            ir.border = S["thin"]
            contrib = ws.cell(r, 5, f"=IF(OR(B{r}=\"\",C{r}=\"\"),\"—\",B{r}*C{r})")
            contrib.number_format = _FMT_SH
            contrib.border = S["thin"]
            r += 1
        last = r - 1
        ws.cell(r, 1, "Expected value").font = S["font_bold"]
        ws.cell(r, 1).border = S["thin"]
        ev = ws.cell(r, 5, f"=SUM(E{first_data}:E{last})")
        ev.font = S["font_kpi"]
        ev.number_format = _FMT_SH
        ev.fill = S["pale_gold"]
        ev.border = S["thin"]
        for col in range(1, 6):
            ws.cell(r, col).border = S["thin"]
            ws.cell(r, col).fill = S["pale_gold"]
        r += 2

    ws["A20"].value = "Last price (for implied return)"
    ws["A20"].font = S["font_muted"]
    ws["B20"].value = price
    ws["B20"].number_format = _FMT_SH
    ws["B20"].fill = S["input_fill"]
    ws["A21"].value = "Report 12-month PT"
    ws["A21"].font = S["font_muted"]
    ws["B21"].value = pt
    ws["B21"].number_format = _FMT_SH

    ws.column_dimensions["A"].width = 36
    for ch in "BCDE":
        ws.column_dimensions[ch].width = 18
    _print_setup(ws, landscape=False)
    ws.sheet_properties.tabColor = SECTION


def _street_sheet(wb, street, S):
    ws = wb.create_sheet("Street")
    _write_banner(ws, 8, "STREET CONSENSUS", "Sell-side ratings and 12-month targets from the report", S)
    if street:
        _dump_md_table(ws, 4, street.get("title") or "ANALYST RATINGS", street, S)
    else:
        ws["A4"] = "No firm-level consensus table found in the saved report."
        ws["A4"].font = S["font_muted"]
    ws.column_dimensions["A"].width = 28
    for ch in "BCDEFGH":
        ws.column_dimensions[ch].width = 16
    _print_setup(ws, landscape=True)
    ws.sheet_properties.tabColor = "1E3A8A"


def _quarterly_sheet(wb, financials: dict, S):
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Quarterly")
    q = (financials or {}).get("quarterly") or {}
    # Prefer explicit current / prior; also list from nested if present
    periods = []
    for key in ("current", "prior_year_same_q"):
        if isinstance(q.get(key), dict):
            periods.append(q[key])
    # If extract_financials_from_db only has those two, that's fine.
    # Also surface YTD.
    ytd_cols = []
    for key, lab in (("current_ytd", "YTD"), ("prior_ytd", "Prior YTD")):
        if isinstance(q.get(key), dict):
            ytd_cols.append((lab, q[key]))

    _write_banner(ws, 8, "QUARTERLY RESULTS", "Latest reported quarter vs year-ago  ·  $ millions", S)
    headers = ["Metric"]
    data_cols = []
    if periods:
        cur = periods[0]
        headers.append(_q_label(cur, "Latest Q"))
        data_cols.append(cur)
        if len(periods) > 1:
            headers.append(_q_label(periods[1], "Year-ago Q"))
            data_cols.append(periods[1])
    for lab, p in ytd_cols:
        headers.append(lab)
        data_cols.append(p)

    for i, h in enumerate(headers, start=1):
        c = ws.cell(4, i, h)
        c.font = S["font_h"]
        c.fill = S["navy_fill"]
        c.alignment = S["center"]
        c.border = S["thin"]

    metrics = [
        ("Revenue", "Revenue", "mm"),
        ("Gross profit", "GrossProfit", "mm"),
        ("Operating income", "OperatingIncome", "mm"),
        ("EBITDA", "EBITDA", "mm"),
        ("Net income", "NetIncome", "mm"),
        ("Diluted EPS", "DilutedEPS", "sh"),
        ("Free cash flow", "FreeCashFlow", "mm"),
        ("Operating cash flow", "OperatingCashFlow", "mm"),
        ("Cash", "Cash", "mm"),
        ("Total debt", "TotalDebt", "mm"),
    ]
    r = 5
    for lab, key, kind in metrics:
        ws.cell(r, 1, lab).font = S["font"]
        ws.cell(r, 1).border = S["thin"]
        for i, p in enumerate(data_cols, start=2):
            val = _period_money(p, key, already_millions=False)
            c = ws.cell(r, i, val)
            c.font = S["font"]
            c.border = S["thin"]
            c.alignment = S["right"]
            c.number_format = _fmt_for(kind)
        r += 1

    if not data_cols:
        ws["A5"] = "No quarterly periods in the financials store for this ticker."
        ws["A5"].font = S["font_muted"]

    ws.column_dimensions["A"].width = 28
    for i in range(2, max(len(headers), 2) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    _print_setup(ws, landscape=True)
    ws.sheet_properties.tabColor = MUTED


def _q_label(p: dict, fallback: str) -> str:
    fp = p.get("fp") or ""
    fy = p.get("fy") or ""
    end = str(p.get("end") or "")[:10]
    if fp and fy:
        return f"{fp} FY{fy}"
    return end or fallback


def _capital_from(financials: dict, price: Optional[float]) -> dict[str, Any]:
    annuals = _annuals_oldest_first(financials)
    ttm = (financials or {}).get("ttm") or {}
    latest = ttm if ttm else (annuals[-1] if annuals else {})
    fy_row = annuals[-1] if annuals else latest
    shares = to_model_units("DilutedShares",
                            latest.get("DilutedShares") or latest.get("SharesOutstanding")
                            or fy_row.get("DilutedShares") or fy_row.get("SharesOutstanding"))
    cash = to_model_units("Cash", latest.get("Cash") if latest.get("Cash") is not None else fy_row.get("Cash"))
    debt = to_model_units("TotalDebt", latest.get("TotalDebt") if latest.get("TotalDebt") is not None else fy_row.get("TotalDebt"))
    equity = to_model_units("StockholdersEquity", fy_row.get("StockholdersEquity"))
    rev = to_model_units("Revenue", latest.get("Revenue") if latest.get("Revenue") is not None else fy_row.get("Revenue"))
    ebitda = to_model_units("EBITDA", latest.get("EBITDA") if latest.get("EBITDA") is not None else fy_row.get("EBITDA"))
    fcf = to_model_units("FreeCashFlow", latest.get("FreeCashFlow") if latest.get("FreeCashFlow") is not None else fy_row.get("FreeCashFlow"))
    eps = _f(fy_row.get("DilutedEPS"))
    ni = to_model_units("NetIncome", fy_row.get("NetIncome"))
    if eps is None and ni is not None and shares:
        eps = ni / shares
    mkt = (price * shares) if (price and shares) else None
    net_debt = None
    if debt is not None or cash is not None:
        net_debt = (debt or 0.0) - (cash or 0.0)
    ev = (mkt + (net_debt or 0.0)) if mkt is not None else None
    pe = (price / eps) if (price and eps and eps > 0) else None
    bvps = (equity / shares) if (equity is not None and shares) else None
    pb = (price / bvps) if (price and bvps and bvps > 0) else None
    ev_eb = (ev / ebitda) if (ev is not None and ebitda and ebitda > 0) else None
    ev_sales = (ev / rev) if (ev is not None and rev and rev > 0) else None
    fcf_y = (fcf / mkt) if (fcf is not None and mkt and mkt > 0) else None
    return {
        "price": price,
        "shares": shares,
        "cash": cash,
        "total_debt": debt,
        "net_debt": net_debt,
        "equity": equity,
        "market_cap": mkt,
        "enterprise_value": ev,
        "pe": pe,
        "pb": pb,
        "ev_ebitda": ev_eb,
        "ev_sales": ev_sales,
        "fcf_yield": fcf_y,
        "book_value_ps": bvps,
        "revenue": rev,
        "ebitda": ebitda,
        "fcf": fcf,
    }


def _upside(pt: Optional[float], price: Optional[float]) -> Optional[float]:
    if pt is None or not price:
        return None
    try:
        return pt / float(price) - 1.0
    except (TypeError, ZeroDivisionError):
        return None


# ── Public API ──────────────────────────────────────────────────────────────
def build_ib_model_bytes(
    ticker: str,
    *,
    financials: Optional[dict] = None,
    report_md: str = "",
    summary: Optional[dict] = None,
    quote: Optional[dict] = None,
    engine: str = "",
    entity_name: str = "",
    sector: str = "",
    industry: str = "",
    source: str = "",
    dropbox_note: str = "",
    generated_at: Optional[str] = None,
) -> bytes:
    """Return .xlsx bytes for an IB-style research model."""
    import openpyxl

    tk = (ticker or "").strip().upper()
    financials = financials if isinstance(financials, dict) else {}
    summary = summary if isinstance(summary, dict) else {}
    quote = quote if isinstance(quote, dict) else {}
    md = report_md or ""
    tables = parse_md_tables(md)
    forecasts = extract_forecasts(tables)
    dcf = extract_dcf(md)
    wacc_build = extract_wacc_build(tables)
    ladder = extract_dcf_ladder(tables)
    sensitivity = extract_sensitivity(tables)
    bridge_tbl = extract_bridge_table(tables)
    scenarios = extract_scenarios(md)
    street = extract_street_table(tables)
    comps = extract_comps_table(tables)
    derivation = extract_derivation_table(tables)

    price = _f(quote.get("price")) or _f(summary.get("current_price"))
    pt = _f(summary.get("price_target"))
    rating = summary.get("rating")
    upside = _f(summary.get("upside_pct"))
    if upside is not None and abs(upside) > 2:
        upside = upside / 100.0
    if upside is None:
        upside = _upside(pt, price)

    thesis = (summary.get("thesis") or "")[:900]
    entity = entity_name or financials.get("entity_name") or tk
    as_of = (generated_at or datetime.now(timezone.utc).strftime("%Y-%m-%d"))[:19]
    capital = _capital_from(financials, price)
    cols = _build_columns(financials, forecasts)

    S = _styles()
    wb = openpyxl.Workbook()
    _cover_sheet(
        wb,
        ticker=tk,
        entity=entity,
        sector=sector or summary.get("sector") or "",
        industry=industry,
        as_of=as_of,
        engine=engine,
        rating=rating,
        pt=pt,
        price=price,
        upside=upside,
        thesis=thesis,
        capital=capital,
        source=source or financials.get("source") or "",
        dropbox_note=dropbox_note,
        S=S,
    )
    _model_sheet(wb, cols, S)
    _valuation_sheet(
        wb,
        capital=capital,
        dcf=dcf,
        wacc_build=wacc_build,
        ladder=ladder,
        sensitivity=sensitivity,
        forecasts=forecasts,
        financials=financials,
        derivation=derivation,
        comps=comps,
        bridge_tbl=bridge_tbl,
        pt=pt,
        price=price,
        S=S,
        approaches=extract_valuation_approaches(
            md, last=price, pt=pt, tables=tables, dcf=dcf,
        ),
    )
    _scenarios_sheet(wb, scenarios, price, pt, S)
    _street_sheet(wb, street, S)
    _quarterly_sheet(wb, financials, S)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_ib_model(ticker: str, output_path: Path | str, **kwargs) -> Path:
    raw = build_ib_model_bytes(ticker, **kwargs)
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(raw)
    return path
